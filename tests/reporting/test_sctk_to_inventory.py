# ============================================================================
#  Copyright (c) nexB Inc. http://www.nexb.com/ - All rights reserved.
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#      http://www.apache.org/licenses/LICENSE-2.0
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#
#  SPDX-License-Identifier: Apache-2.0
# ============================================================================

from __future__ import absolute_import
from __future__ import print_function

import os

from click.testing import CliRunner
from openpyxl import load_workbook
import shutil

from commoncode.testcase import FileBasedTesting
from commoncode.testcase import FileDrivenTesting
from reporting import sctk_to_inventory


# These 2 lines are used by the .xlsx test(s) at the bottom, outside of
# class TestScanToInventory(FileBasedTesting)
test_env = FileDrivenTesting()
test_env.test_data_dir = os.path.join(os.path.dirname(__file__), 'data')


class TestScanToInventory(FileBasedTesting):

    test_data_dir = os.path.join(os.path.dirname(__file__), 'data')

    def test_update_data_field(self):
        test_file = self.get_test_loc('sctk2inv/test.json')
        info_list, concluded_package_field_name = sctk_to_inventory.get_data_from_json(
            test_file)
        report = True
        result = sctk_to_inventory.update_data_field(
            info_list, concluded_package_field_name, report)
        expected_list = [{'Resource Path': 'test/project/express_nodeJs', 'path': 'test/project/express_nodeJs',
                          'Item Type': 'D', 'type': 'directory', 'Resource Name': 'express_nodeJs',
                          'name': 'express_nodeJs', 'base_name': 'express_nodeJs', 'extension': '',
                          'size': 0, 'date': None, 'sha1': None, 'md5': None, 'sha256': None, 'mime_type': None,
                          'file_type': None, 'Language': None, 'programming_language': None, 'is_binary': False,
                          'is_text': False, 'is_archive': False, 'is_media': False, 'is_source': False, 'is_script': False,
                          'license_key': '', 'license_score': '', 'license_expressions': '', 'percentage_of_license_text': 0,
                          'copyrights': '', 'holders': '', 'Detected Copyright': '', 'authors': '', 'package_data': '',
                          'for_packages': 'pkg:maven/abc', 'emails': '', 'urls': '', 'files_count': 822, 'dirs_count': 608,
                          'size_count': 4099422, 'scan_errors': ''},
                         {'Resource Path': 'test/project/express_nodeJs/package-lock.json',
                          'path': 'test/project/express_nodeJs/package-lock.json', 'Item Type': 'F', 'type': 'file',
                          'Resource Name': 'package-lock.json', 'name': 'package-lock.json', 'base_name': 'package-lock',
                          'extension': '.json', 'size': 21076, 'date': '2021-11-23', 'sha1': 'aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa',
                          'md5': 'aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa', 'sha256': 'aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa',
                          'mime_type': 'application/json', 'file_type': 'JSON data', 'Language': 'JavaScript',
                          'programming_language': None, 'is_binary': False, 'is_text': True, 'is_archive': False,
                          'is_media': False, 'is_source': False, 'is_script': False, 'license_key': '', 'license_score': '',
                          'license_expressions': '', 'percentage_of_license_text': 0, 'copyrights': '', 'holders': '',
                          'Detected Copyright': '', 'authors': '',
                          'package_manifests': [{'type': 'npm', 'namespace': None, 'name': 'send', 'version': '0.16.2',
                                                 'qualifiers': {}, 'subpath': None, 'primary_language': 'JavaScript',
                                                 'description': None, 'release_date': None, 'parties': [], 'keywords': [],
                                                 'homepage_url': None, 'download_url': 'https://registry.npmjs.org/send/-/send-0.16.2.tgz',
                                                 'size': None, 'sha1': None, 'md5': None, 'sha256': None, 'sha512': None,
                                                 'bug_tracking_url': None, 'code_view_url': None, 'vcs_url': None, 'copyright': None,
                                                 'license_expression': None, 'declared_license': None, 'notice_text': None,
                                                 'root_path': 'test/project/express_nodeJs',
                                                 'dependencies': 'pkg:npm/debug@2.6.9\npkg:npm/depd@%7E1.1.2',
                                                 'contains_source_code': None, 'source_packages': [], 'extra_data': {},
                                                 'purl': 'pkg:npm/send@0.16.2', 'repository_homepage_url': 'https://www.npmjs.com/package/send',
                                                 'repository_download_url': 'https://registry.npmjs.org/send/-/send-0.16.2.tgz',
                                                 'api_data_url': 'https://registry.npmjs.org/send/0.16.2'}],
                          'package__type': 'npm', 'package__namespace': '', 'package__name': 'send', 'package__version': '0.16.2',
                          'package__qualifiers': '', 'package__subpath': '', 'package__primary_language': 'JavaScript',
                          'package__description': '', 'package__release_date': '', 'package__parties': '', 'package__keywords': '',
                          'package__homepage_url': '', 'package__download_url': 'https://registry.npmjs.org/send/-/send-0.16.2.tgz',
                          'package__size': '', 'package__sha1': '', 'package__md5': '', 'package__sha256': '', 'package__sha512': '',
                          'package__bug_tracking_url': '', 'package__code_view_url': '', 'package__vcs_url': '',
                          'package__copyright': '', 'package__license_expression': '', 'package__declared_license': '',
                          'package__notice_text': '', 'package__root_path': 'test/project/express_nodeJs',
                          'package__dependencies': 'pkg:npm/debug@2.6.9\npkg:npm/depd@%7E1.1.2', 'package__contains_source_code': '',
                          'package__source_packages': '', 'package__extra_data': '', 'package__purl': 'pkg:npm/send@0.16.2',
                          'package__repository_homepage_url': 'https://www.npmjs.com/package/send',
                          'package__repository_download_url': 'https://registry.npmjs.org/send/-/send-0.16.2.tgz',
                          'package__api_data_url': 'https://registry.npmjs.org/send/0.16.2', 'emails': '', 'urls': '',
                          'files_count': 0, 'dirs_count': 0, 'size_count': 0, 'scan_errors': ''}]
        assert result == expected_list

    def test_data_str_convertion(self):
        test1 = ['123', '456']
        test2 = '123'
        test3 = {'123': '456'}
        test4 = [{'123': '456'}]
        test5 = {'123': ['456']}

        result1 = sctk_to_inventory.data_str_convertion(test1)
        result2 = sctk_to_inventory.data_str_convertion(test2)
        result3 = sctk_to_inventory.data_str_convertion(test3)
        result4 = sctk_to_inventory.data_str_convertion(test4)
        result5 = sctk_to_inventory.data_str_convertion(test5)

        assert result1 == '123\n456'
        assert result2 == '123'
        assert result3 == '123: 456'
        assert result4 == '123: 456'
        assert result5 == '123: 456'

    def test_get_data_from_json_package_manifests(self):
        test_file = self.get_test_loc('sctk2inv/test.json')
        info_list, concluded_package_field_name = sctk_to_inventory.get_data_from_json(
            test_file)

        assert concluded_package_field_name == 'package_manifests'
        assert concluded_package_field_name != 'package_data'
        assert concluded_package_field_name != 'packages'
        assert concluded_package_field_name is not None

    def test_get_data_from_json_package_data(self):
        test_file = self.get_test_loc(
            'sctk2inv/setup-py-clip-file-cat-2022-05-09-01.json')
        info_list, concluded_package_field_name = sctk_to_inventory.get_data_from_json(
            test_file)

        assert concluded_package_field_name != 'package_manifests'
        assert concluded_package_field_name == 'package_data'
        assert concluded_package_field_name != 'packages'
        assert concluded_package_field_name is not None

    def test_get_data_from_json_None(self):
        test_file = self.get_test_loc(
            'sctk2inv/setup-py-info-file-cat-2022-05-09-01.json')
        info_list, concluded_package_field_name = sctk_to_inventory.get_data_from_json(
            test_file)

        assert concluded_package_field_name != 'package_manifests'
        assert concluded_package_field_name != 'package_data'
        assert concluded_package_field_name != 'packages'
        assert concluded_package_field_name is None


def test_sctkv31_info_json():
    input_json = test_env.get_test_loc(
        'sctk2inv/setup-py-info-file-cat-2022-05-09-01.json')
    expected_result_file = test_env.get_test_loc(
        'sctk2inv/setup-py-info-file-cat-2022-05-09-01-scan2inv.xlsx')
    actual_result_file = test_env.get_temp_file('test-out.xlsx')
    report = True
    options = ['--report', input_json, actual_result_file]
    runner = CliRunner()
    _ = runner.invoke(sctk_to_inventory.cli, options, catch_exceptions=False)
    check_results(actual_result_file, expected_result_file, regen=False)


def test_sctkv31_clip_json():
    input_json = test_env.get_test_loc(
        'sctk2inv/setup-py-clip-file-cat-2022-05-09-01.json')
    expected_result_file = test_env.get_test_loc(
        'sctk2inv/setup-py-clip-file-cat-2022-05-09-01-scan2inv.xlsx')
    actual_result_file = test_env.get_temp_file('test-out.xlsx')
    options = ['--report', input_json, actual_result_file]
    runner = CliRunner()
    _ = runner.invoke(sctk_to_inventory.cli, options, catch_exceptions=False)
    check_results(actual_result_file, expected_result_file, regen=False)


def check_results(actual_result_file, expected_result_file, regen=False):
    """
    Compare actual result file with expected result file.
    """
    result_wb = load_workbook(actual_result_file)
    if regen:
        # We overwrite `expected_result_file` with the contents of `actual_result_file`
        # for the purpose of regenerating test files
        shutil.copy2(actual_result_file, expected_result_file)
    expected_wb = load_workbook(expected_result_file)

    # Load named worksheets from workbooks
    results = result_wb['INVENTORY']
    expected = expected_wb['INVENTORY']

    # Look at each row from both worksheets at the same time
    for result_row, expected_row in zip(results.iter_rows(), expected.iter_rows()):
        # Look at each cell from the row we are looking at from above
        for result_cell, expected_cell in zip(result_row, expected_row):
            # Make sure we have the same value in each cell
            assert result_cell.value == expected_cell.value

    # compare the number of rows in the 2 sheets
    results_rows = results.max_row
    expected_rows = expected.max_row
    assert results_rows == expected_rows

    # This is another way to compare the row count but does not use the openpyxl 'max_row' method.
    results_count = 0
    for row in results:
        if not all([cell.value == None for cell in row]):
            results_count += 1

    expected_count = 0
    for row in expected:
        if not all([cell.value == None for cell in row]):
            expected_count += 1

    assert results_count == expected_count
