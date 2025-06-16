# ============================================================================
#  Copyright (c) nexB Inc. http://www.nexb.com/ - All rights reserved.
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#      http://www.apache.org/licenses/LICENSE-2.0
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#
#  SPDX-License-Identifier: Apache-2.0
# ============================================================================

from __future__ import absolute_import
from __future__ import print_function

from collections import namedtuple
import json
import os

from openpyxl import load_workbook
import shutil

from commoncode.testcase import FileBasedTesting

from reporting.license_category import add_data_to_expression
# from reporting.license_category import check_parse_errors
# from reporting.license_category import check_dejacode_existance
from reporting.license_category import collect_expression_info
from reporting.license_category import deduplicate_and_simplify
from reporting.license_category import process_output_bom
from reporting.license_category import parse_expression
# from reporting.license_category import report_errors
from reporting.license_category import set_with_exception_as_primary
from reporting.license_category import validate_license_expression
from utilitycode.bom_utils import create_xlsx_output


class TestLicenseCategory(FileBasedTesting):

    test_data_dir = os.path.join(os.path.dirname(__file__), "data")

    def test_invalid_license_multi_missing(self):
        test_file = self.get_test_loc(
            "license_category/check-dje-existance-multi-missing.json"
        )

        LicenseExpressionInfo = namedtuple(
            "LicenseExpressionInfo", ["coord", "key"])

        infos = [
            LicenseExpressionInfo(coord="A2", key="mit"),
            LicenseExpressionInfo(coord="A3", key="foo"),
            LicenseExpressionInfo(coord="A4", key="mit"),
            LicenseExpressionInfo(coord="A5", key="bar"),
        ]

        dje_data = json.loads(open(test_file).read())
        errors = validate_license_expression(infos, dje_data)
        expected_error = ['Invalid license_expression value: "foo" at coord: "A3"',
                          'Invalid license_expression value: "bar" at coord: "A5"']
        assert errors == expected_error

    def test_license_no_errors(self):
        test_file = self.get_test_loc(
            "license_category/check-dje-existance-no-errors.json"
        )

        LicenseExpressionInfo = namedtuple(
            "LicenseExpressionInfo", ["coord", "key"])

        infos = [
            LicenseExpressionInfo(coord="A2", key="mit"),
            LicenseExpressionInfo(coord="A3", key="apache-2.0"),
        ]

        dje_data = json.loads(open(test_file).read())
        potential_errors = validate_license_expression(infos, dje_data)

        assert not potential_errors

    def test_collect_expression_info_dual_key_expression_missing(self):
        test_file = self.get_test_loc(
            "license_category/collect-expression-info-dual-key-expression-missing.xlsx"
        )
        test_bom = load_workbook(test_file)
        test_ws = test_bom.active

        result, errors = collect_expression_info(test_ws, "A")
        expected_error = ['Empty license_expression value at A6']
        assert errors == expected_error

        result_coords = [le.coord for le in result]
        result_keys = [le.key for le in result]

        assert result_coords == [
            "A2",
            "A2",
            "A3",
            "A3",
            "A4",
            "A4",
            "A5",
            "A5",
            "A7",
            "A7",
        ]
        assert result_keys == [
            "mit",
            "gpl-2.0",
            "apache-2.0",
            "bsd-new",
            "agpl-3.0",
            "public-domain",
            "apache-2.0",
            "mit",
            "apache-2.0",
            "mit",
        ]

    def test_collect_expression_info_dual_key_expression(self):
        test_file = self.get_test_loc(
            "license_category/collect-expression-info-dual-key-expression.xlsx"
        )
        test_bom = load_workbook(test_file)
        test_ws = test_bom.active

        result, _err = collect_expression_info(test_ws, "A")

        result_coords = [le.coord for le in result]
        result_keys = [le.key for le in result]

        assert result_coords == ["A2", "A2",
                                 "A3", "A3", "A4", "A4", "A5", "A5"]
        assert result_keys == [
            "mit",
            "gpl-2.0",
            "apache-2.0",
            "bsd-new",
            "agpl-3.0",
            "public-domain",
            "apache-2.0",
            "mit",
        ]

    def test_collect_expression_info_single_key_expression(self):
        test_file = self.get_test_loc(
            "license_category/collect-expression-info-single-key-expression.xlsx"
        )
        test_bom = load_workbook(test_file)
        test_ws = test_bom.active

        result, _err = collect_expression_info(test_ws, "A")

        result_coords = [le.coord for le in result]
        result_keys = [le.key for le in result]

        assert result_coords == ["A2", "A3", "A4", "A5"]
        assert result_keys == ["mit", "apache-2.0", "gpl-2.0", "apache-2.0"]

    def test_collect_expression_info_single_key_expression(self):
        test_file = self.get_test_loc(
            "license_category/collect-expression-info-single-key-expression.xlsx"
        )
        test_bom = load_workbook(test_file)
        test_ws = test_bom.active

        result, _err = collect_expression_info(test_ws, "A")

        result_coords = [le.coord for le in result]
        result_keys = [le.key for le in result]

        assert result_coords == ["A2", "A3", "A4", "A5"]
        assert result_keys == ["mit", "apache-2.0", "gpl-2.0", "apache-2.0"]

    def test_collect_expression_info1(self):
        test_file = self.get_test_loc(
            "license_category/check-parse-errors-multi-error.xlsx"
        )
        test_bom = load_workbook(test_file)

        worksheet = test_bom.active
        _info, errors = collect_expression_info(worksheet, "A")
        expected_err = ['Empty license_expression value at A4', 'WARNING: Failed to decode license_expression: and at: A6',
                        'WARNING: Failed to decode license_expression: mit OR at: A7']
        assert len(errors) == 3
        assert errors == expected_err

    def test_validate_license_expression(self):
        test_file = self.get_test_loc(
            "license_category/check-parse-errors-expression-error.xlsx"
        )
        test_bom = load_workbook(test_file)
        tmp_lic_dict = {'mit': {}}
        worksheet = test_bom.active
        info, _errors = collect_expression_info(worksheet, "A")
        err = validate_license_expression(info, tmp_lic_dict)
        expected_err = ['Invalid license_expression value: "apache-2.0 AAND bsd-new" at coord: "A3"',
                        'Invalid license_expression value: "invalid_license" at coord: "A4"']
        assert len(err) == 2
        assert err == expected_err

    def test_no_errors(self):
        test_file = self.get_test_loc(
            "license_category/check-parse-errors-no-errors.xlsx"
        )
        test_bom = load_workbook(test_file)
        test_bom = load_workbook(test_file)
        tmp_lic_dict = {'mit': {}, 'gpl-2.0': {},
                        'public-domain': {}, 'apache-2.0': {}, 'lgpl-2.0-plus': {}}
        worksheet = test_bom.active
        info, errors = collect_expression_info(worksheet, "A")
        err = validate_license_expression(info, tmp_lic_dict)

        assert not errors
        assert not err

    def test_parse_expression_multi_and(self):
        test_coordinate = 'B2'
        result1, err1 = parse_expression("mit AND apache-2.0", test_coordinate)
        result2, err2 = parse_expression(
            "mit AND apache-2.0 AND gpl-2.0", test_coordinate)
        result3, err3 = parse_expression(
            "mit AND (apache-2.0 AND apache-2.0)", test_coordinate)
        assert set(result1) == set(["mit", "apache-2.0"])
        assert set(result2) == set(["mit", "apache-2.0", "gpl-2.0"])
        assert set(result3) == set(["mit", "apache-2.0"])
        assert not err1
        assert not err2
        assert not err3

    def test_parse_expression_multi_or(self):
        test_coordinate = 'B2'
        result1, err1 = parse_expression("mit OR apache-2.0", test_coordinate)
        result2, err2 = parse_expression(
            "mit OR apache-2.0 OR gpl-2.0", test_coordinate)
        result3, err3 = parse_expression(
            "mit OR (apache-2.0 OR apache-2.0)", test_coordinate)
        assert set(result1) == set(["mit", "apache-2.0"])
        assert set(result2) == set(["mit", "apache-2.0", "gpl-2.0"])
        assert set(result3) == set(["mit", "apache-2.0"])
        assert not err1
        assert not err2
        assert not err3

    def test_parse_expression_single(self):
        test_coordinate = 'B2'
        result, err = parse_expression("mit", test_coordinate)
        assert result == ["mit"]
        assert not err

    def test_parse_expression_none(self):
        test_coordinate = 'B2'
        result, err = parse_expression("", test_coordinate)
        assert not result
        assert not err

    def test_create_expression_with_license_data_single(self):
        license_expression = "mit"
        license_data = {
            "mit": {
                "short_name": "MIT License",
                "is_exception": False,
                "category": "Permissive",
                "attribution_required": True,
                "redistribution_required": False,
            }
        }
        result = add_data_to_expression(
            license_expression, license_data, "")
        assert result.render() == "mit"
        assert result.render(template="{symbol.wrapped.name}") == "MIT License"
        assert result.render(
            template="{symbol.wrapped.category}") == "Permissive"

    def test_create_expression_with_license_data_OR(self):
        license_expression = "mit OR bsd-new"
        license_data = {
            "mit": {
                "short_name": "MIT License",
                "is_exception": False,
                "category": "Permissive",
                "attribution_required": True,
                "redistribution_required": False,
            },
            "bsd-new": {
                "short_name": "BSD-3-Clause",
                "is_exception": False,
                "category": "Permissive",
                "attribution_required": True,
                "redistribution_required": False,
            },
        }
        result = add_data_to_expression(
            license_expression, license_data, "")
        assert result.render() == "mit OR bsd-new"
        assert (
            result.render(template="{symbol.wrapped.name}")
            == "MIT License OR BSD-3-Clause"
        )
        assert (
            result.render(template="{symbol.wrapped.category}")
            == "Permissive OR Permissive"
        )

    def test_create_expression_with_license_data_OR_AND(self):
        license_expression = "mit OR (bsd-new AND gpl-2.0)"
        license_data = {
            "mit": {
                "short_name": "MIT License",
                "is_exception": False,
                "category": "Permissive",
                "attribution_required": True,
                "redistribution_required": False,
            },
            "bsd-new": {
                "short_name": "BSD-3-Clause",
                "is_exception": False,
                "category": "Permissive",
                "attribution_required": True,
                "redistribution_required": False,
            },
            "gpl-2.0": {
                "short_name": "GPL 2.0",
                "is_exception": False,
                "category": "Copyleft",
                "attribution_required": True,
                "redistribution_required": False,
            },
        }
        result = add_data_to_expression(
            license_expression, license_data, "")
        assert result.render() == "mit OR (bsd-new AND gpl-2.0)"
        assert (
            result.render(template="{symbol.wrapped.name}")
            == "MIT License OR (BSD-3-Clause AND GPL 2.0)"
        )
        assert (
            result.render(template="{symbol.wrapped.category}")
            == "Permissive OR (Permissive AND Copyleft)"
        )

    def test_set_with_exception_as_primary(self):
        exp1 = "Permissive AND Copyleft WITH Copyleft Limited"
        expected1 = "Permissive AND Copyleft Limited"

        exp2 = "(Permissive OR Copyleft WITH Copyleft Limited) OR Permissive"
        expected2 = "(Permissive OR Copyleft Limited) OR Permissive"

        exp3 = "(bsd-new AND gpl-2.0 WITH classpath-exception) OR mit"
        expected3 = "(bsd-new AND classpath-exception) OR mit"

        result1 = set_with_exception_as_primary(exp1)
        result2 = set_with_exception_as_primary(exp2)
        result3 = set_with_exception_as_primary(exp3)
        assert result1 == expected1
        assert result2 == expected2
        assert result3 == expected3

    def test_process_output_bom(self):
        """
        This function tests the use of the 4 new flags (-lc, -ln, -a and -r) to prevent overwriting a
        row's values in those columns when the auditor has already provided a value in the License
        Name column for that row.
        Load the .json file containing the output from an API call made by
        1.  uncommenting the json.dump() code inside the cli() function in license_category.py and
        2.  running (following should be a single line):
                lcat -e P -o Q -lc L -ln M -a N -r O
                /home/tests/reporting/data/license_category/lcat-no-row-overwrite-input.xlsx
                /home/tests/reporting/data/license_category/lcat-no-row-overwrite-output.xlsx
                [DejaCode API key]
        """
        api_data_file01 = self.get_test_loc(
            "license_category/dejacode-api-call-20211209.json"
        )
        license_data = json.load(open(api_data_file01))

        test_input_file = self.get_test_loc(
            "license_category/lcat-no-row-overwrite-input.xlsx"
        )
        output_xlsx = self.get_temp_file("test-out.xlsx")
        expected_xlsx = self.get_test_loc(
            "license_category/lcat-no-row-overwrite-output.xlsx"
        )
        input_bom = load_workbook(test_input_file)
        input_ws = input_bom.active

        result = process_output_bom(
            input_ws,
            "P",
            "Q",
            "L",
            "M",
            "N",
            "O",
            license_data
        )
        create_xlsx_output(output_xlsx, result)
        check_results(output_xlsx, expected_xlsx, regen=False)

    def test_process_output_bom_more_recent(self):
        """
        This is a copy of test_create_output_bom() above but using a more recent API
        call/json.dump and different input/output files to retrieve current data from DejaCode.
        """
        api_data_file01 = self.get_test_loc(
            "license_category/dejacode-api-call-20220620.json"
        )
        license_data = json.load(open(api_data_file01))

        test_input_file = self.get_test_loc(
            "license_category/check-cli-errors-multi-error-more-overwrite-testing.xlsx"
        )
        output_xlsx = self.get_temp_file("test-out.xlsx")
        expected_xlsx = self.get_test_loc(
            "license_category/check-cli-errors-multi-error-more-overwrite-testing-lcat-2022-06-20-overwrite-force.xlsx"
        )
        input_bom = load_workbook(test_input_file)
        input_ws = input_bom.active

        result = process_output_bom(
            input_ws,
            "Y",
            "Z",
            "U",
            "V",
            "W",
            "X",
            license_data
        )
        create_xlsx_output(output_xlsx, result)
        check_results(output_xlsx, expected_xlsx, regen=False)

    def test_deduplicate_and_simplify(self):
        cat1 = "((Copyleft AND Copyleft AND Copyleft) OR Permissive) AND (Permissive AND Permissive)"
        cat2 = "Permissive"
        cat3 = "Permissive AND Permissive"
        cat4 = "Permissive OR Permissive"
        cat5 = "Permissive AND Copyleft Limited AND Permissive"

        expected1 = "(Copyleft OR Permissive) AND Permissive"
        expected2 = "Permissive"
        expected3 = "Permissive"
        expected4 = "Permissive"
        expected5 = "Permissive AND Copyleft Limited"

        result1 = deduplicate_and_simplify(cat1)
        result2 = deduplicate_and_simplify(cat2)
        result3 = deduplicate_and_simplify(cat3)
        result4 = deduplicate_and_simplify(cat4)
        result5 = deduplicate_and_simplify(cat5)

        assert expected1 == result1
        assert expected2 == result2
        assert expected3 == result3
        assert expected4 == result4
        assert expected5 == result5


"""
def test_report_errors(caplog, capsys):
    errors = [
        'DejaCode Empty license_expression value at coord: "Y5"',
        'DejaCode Empty license_expression value at coord: "Y6"',
        'DejaCode Missing Data error: license_key value: "foo-3.0" at coord: "Y2"',
        'DejaCode Missing Data error: license_key value: "commercial" at coord: "Y9"',
        'DejaCode Missing Data error: license_key value: "Acme_EULA.txt" at coord: "Y10"',
    ]

    force = True

    report_errors(errors, force)

    out, err = capsys.readouterr()

    # test log
    assert 'DejaCode Empty license_expression value at coord: "Y5"' in caplog.text
    assert 'DejaCode Empty license_expression value at coord: "Y6"' in caplog.text
    assert (
        'DejaCode Missing Data error: license_key value: "foo-3.0" at coord: "Y2"'
        in caplog.text
    )
    assert (
        'DejaCode Missing Data error: license_key value: "commercial" at coord: "Y9"'
        in caplog.text
    )
    assert (
        'DejaCode Missing Data error: license_key value: "Acme_EULA.txt" at coord: "Y10"'
        in caplog.text
    )

    # test stderr
    assert 'DejaCode Empty license_expression value at coord: "Y5"' in err
    assert 'DejaCode Empty license_expression value at coord: "Y6"' in err
    assert (
        'DejaCode Missing Data error: license_key value: "foo-3.0" at coord: "Y2"'
        in err
    )
    assert (
        'DejaCode Missing Data error: license_key value: "commercial" at coord: "Y9"'
        in err
    )
    assert (
        'DejaCode Missing Data error: license_key value: "Acme_EULA.txt" at coord: "Y10"'
        in err
    )

    # test stdout
    assert 'DejaCode Empty license_expression value at coord: "Y5"' not in out
"""


def check_results(result_file, expected_file, regen=False):
    """
    Check to see if the contents of the "INVENTORY" sheet from the XLSX files
    `result_file` and `expected_file` are the same.
    If `regen` is True, then overrite `expected_file` with the contents of
    `result_file`
    """
    # result_wb = openpyxl.load_workbook(result_file)
    result_wb = load_workbook(result_file)
    if regen:
        # We overwrite `expected_file` with the contents of `result_file`
        # for the purpose of regenertating test files
        shutil.copy2(result_file, expected_file)
    # expected_wb = openpyxl.load_workbook(expected_file)
    expected_wb = load_workbook(expected_file)

    # Load "Sheet" worksheet from workbooks (modify if necessary for your
    # files' worksheet names)
    results = result_wb["Sheet"]
    expected = expected_wb["Sheet"]

    # Look at each row from both worksheets at the same time
    for result_row, expected_row in zip(results.iter_rows(), expected.iter_rows()):
        # Look at each cell from the row we are looking at from above
        for result_cell, expected_cell in zip(result_row, expected_row):
            # Make sure we have the same value in each cell
            assert result_cell.value == expected_cell.value
