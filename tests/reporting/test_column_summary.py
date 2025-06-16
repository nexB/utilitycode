# ============================================================================
#  Copyright (c) nexB Inc. http://www.nexb.com/ - All rights reserved.
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#      http://www.apache.org/licenses/LICENSE-2.0
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#
#  SPDX-License-Identifier: Apache-2.0
# ============================================================================

from __future__ import absolute_import, print_function

from collections import Counter
from openpyxl import load_workbook
import os

from commoncode.testcase import FileBasedTesting
from reporting import column_summary


class TestComponentSummary(FileBasedTesting):

    test_data_dir = os.path.join(os.path.dirname(__file__), 'data')

    def test_component_rows_multi(self):
        test_bom = load_workbook(self.get_test_loc('column_summary/component-rows-multi.xlsx'))

        test_components_counter = Counter(
            [('component1', 'version1',), ('component2', 'version2',), ('component3', 'version3',)]
        )

        result = {k:v for k,v in column_summary.component_rows(test_bom.active, test_components_counter, ['A','B'])}

        assert result[('component1', 'version1',)] == [1, 'component1', 'version1', 'mit']
        assert result[('component2', 'version2',)] == [1, 'component2', 'version2', 'gpl-2.0']
        assert result[('component3', 'version3',)] == [1, 'component3', 'version3', 'apache-2.0']

    def test_component_rows_single(self):
        test_bom = load_workbook(self.get_test_loc('column_summary/component-rows-single.xlsx'))

        test_components_counter = Counter([('component1',), ('component2',), ('component3',)])

        result = {k:v for k,v in column_summary.component_rows(test_bom.active, test_components_counter, ['A'])}

        assert result[('component1',)] == [1, 'component1', 'version1', 'mit']
        assert result[('component2',)] == [1, 'component2', 'version2', 'gpl-2.0']
        assert result[('component3',)] == [1, 'component3', 'version3', 'apache-2.0']

    def test_components_multi(self):
        test_bom = load_workbook(self.get_test_loc('column_summary/components-multi.xlsx'))

        result = list(column_summary.components(test_bom.active, ['A', 'B']))

        assert result[0] == ('component1', 'version1',)
        assert result[1] == ('component2', 'version2',)
        assert result[2] == ('component3', 'version3',)

    def test_components_single(self):
        test_bom = load_workbook(self.get_test_loc('column_summary/components-single.xlsx'))

        result = list(column_summary.components(test_bom.active, ['A']))

        assert result[0] == ('component1',)
        assert result[1] == ('component2',)
        assert result[2] == ('component3',)
