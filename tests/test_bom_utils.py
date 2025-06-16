# ============================================================================
#  Copyright (c) nexB Inc. http://www.nexb.com/ - All rights reserved.
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#      http://www.apache.org/licenses/LICENSE-2.0
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#
#  SPDX-License-Identifier: Apache-2.0
# ============================================================================

from __future__ import absolute_import, print_function

import os

import openpyxl
from openpyxl import load_workbook

from commoncode.testcase import FileBasedTesting
from utilitycode import bom_utils
from reporting.license_category import collect_expression_info


class TestBomUtils(FileBasedTesting):

    test_data_dir = os.path.join(os.path.dirname(__file__), 'data')

    def test_add_nexb_bom_style(self):
        data = [['header 1', 'header 2', 'header 3'],
                ['content 1', 'content 2', 'content 3'],
                ['content 1', 'content 2', 'content 3'],
                ['content 1', 'content 2', 'content 3']]

        result_wb = bom_utils.create_nexb_bom(data)
        ws = result_wb.active

        for row in ws.iter_rows():
            for cell in row:
                if row == list(ws.iter_rows())[0]:
                    assert cell.font == openpyxl.styles.fonts.Font(
                        name='Arial', size=10, b=True)
                    assert cell.fill == openpyxl.styles.fills.PatternFill(
                        'solid', fgColor='c5ffff')
                else:
                    assert cell.font == openpyxl.styles.fonts.Font(
                        name='Arial', size=10)
                    assert cell.alignment == openpyxl.styles.alignment.Alignment(
                        vertical='center', wrapText=True)

    def test_remove_version_char(self):
        test_vals = ['v 2.1.3', 'v 2015_bleh', 'v 2.0', 'v 2']
        expected = ['2.1.3', '2015_bleh', '2.0', '2']
        for test_val, expected_val in zip(test_vals, expected):
            result = bom_utils.remove_version_char(test_val)
            assert result == expected_val

    def test_get_column(self):
        from openpyxl import Workbook
        ws = load_workbook(self.get_test_loc(
            'bom_utils/get_column.xlsx'), read_only=False).active
        column = bom_utils.get_column(ws, 'A')
        for each in column:
            assert each.value == 'first' or 'one'

        column = bom_utils.get_column(ws, 'B')
        for each in column:
            assert each.value == 'second' or 'two'

        column = bom_utils.get_column(ws, 'F')
        assert column is None

    def test_count_headers(self):
        from openpyxl import load_workbook
        wb = load_workbook(self.get_test_loc(
            'bom_utils/count_headers.xlsx'), read_only=True)
        input_list = [wb, wb]
        result = bom_utils.count_headers(input_list)
        expected = [('Header1', 8), ('Header2', 8),
                    ('Header3', 8), ('Header4', 8)]
        assert set(result) == set(expected)

    def test_count_sheets(self):
        from openpyxl import load_workbook
        wb = load_workbook(self.get_test_loc(
            'bom_utils/count_sheets.xlsx'), read_only=True)
        input_list = [wb, wb, wb, wb]
        result = bom_utils.count_sheets(input_list)
        expected = [('Sheet1', 4), ('Sheet2', 4), ('Sheet3', 4), ('Sheet4', 4)]
        assert set(result) == set(expected)

    def test_get_headers(self):
        from openpyxl import load_workbook
        workbook = load_workbook(self.get_test_loc(
            'bom_utils/headers.xlsx'), read_only=True)
        worksheet = workbook['Sheet1']
        expected_headers = ['Header1', 'Header2', 'Header3', 'Header4']
        result = bom_utils.get_headers(worksheet)
        assert set(result) == set(expected_headers)

    def test_curate_value(self):
        import time
        import datetime
        test_vals = [
            datetime.datetime(2016, 1, 15, 13, 11, 9,
                              103124), ' lead', 'trail ',
            ' both ', 'neither', 1, 2.2, '4.3.2',
        ]
        expected_vals = [
            u'2016-01-15 13:11:09.103124', u'lead', u'trail', u'both', u'neither', u'1',
            u'2.2', u'4.3.2',
        ]
        for test, expected in zip(test_vals, expected_vals):
            assert bom_utils.curate_value(test) == expected

    def test_has_macros(self):
        macros_wb = load_workbook(self.get_test_loc(
            'bom_utils/contains_macros.xlsx'), read_only=True)
        no_macros_wb = load_workbook(self.get_test_loc(
            'bom_utils/no_macros.xlsx'), read_only=True)
        assert bom_utils.has_macros(macros_wb) == [
            'A2', 'B2', 'C2', 'D2', 'A3']
        assert bom_utils.has_macros(no_macros_wb) == []

    def test_load_xlsx_files(self):
        multi = [
            self.get_test_loc('bom_utils/load/xlsx_load1.xlsx'),
            self.get_test_loc('bom_utils/load/xlsx_load2.xlsx'),
            self.get_test_loc('bom_utils/load/xlsx_load3.xlsx'),
        ]
        result = []
        loaded_single = bom_utils.load_xlsx_files([multi[0]])
        loaded_multi = bom_utils.load_xlsx_files(multi)
        expected_file_name = ['xlsx_load1.xlsx']
        expected_file_names = ['xlsx_load1.xlsx',
                               'xlsx_load2.xlsx', 'xlsx_load3.xlsx']
        assert set(expected_file_name) == set(loaded_single.keys())
        assert set(expected_file_names) == set(loaded_multi.keys())

    def test_get_xlsx_paths(self):
        file_1 = self.get_test_loc('bom_utils/paths/xlsx_path1.xlsx')
        file_2 = self.get_test_loc('bom_utils/paths/xlsx_path2.xlsx')
        file_3 = self.get_test_loc('bom_utils/paths/xlsx_path3.xlsx')
        dir_1 = self.get_test_loc('bom_utils/paths')
        result_file = bom_utils.get_xlsx_paths(file_1)
        result_dir = bom_utils.get_xlsx_paths(dir_1)
        expected_file = [file_1]
        expected_dir = [file_1, file_2, file_3]
        assert result_file == expected_file
        assert sorted(result_dir) == expected_dir

    def test_get_header_fields_and_index_dict(self):
        location = self.get_test_loc('bom_utils/input.xlsx')
        input_bom = load_workbook(location)
        input_ws = input_bom.active
        header_dict = bom_utils.get_header_fields_and_index_dict(input_ws)
        expected = {'Resource Path': 'A', 'Resource Name': 'B',
                    'Concluded License Expression': 'C', 'homepage_url': 'D', 'license_url': 'E'}
        assert header_dict == expected

    def test_get_lic_key_and_coord(self):
        location = self.get_test_loc('bom_utils/input.xlsx')
        bom = load_workbook(location)
        ws = bom.active
        expression_info, _err = collect_expression_info(ws, 'C')
        expected = {'mit': ['C2', 'C4'], 'bsd-n': ['C3']}
        result = bom_utils.get_lic_key_and_coord(expression_info)
        assert result == expected
