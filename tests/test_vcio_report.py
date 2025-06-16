# ============================================================================
#  Copyright (c) nexB Inc. http://www.nexb.com/ - All rights reserved.
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#      http://www.apache.org/licenses/LICENSE-2.0
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#
#  SPDX-License-Identifier: Apache-2.0
# ============================================================================

import json
import os
import shutil
from unittest import mock

import openpyxl
from click.testing import CliRunner
from commoncode.testcase import FileDrivenTesting

from vcio import vcio_report

test_env = FileDrivenTesting()
test_env.test_data_dir = os.path.join(os.path.dirname(__file__), "data")


class TestVcioReport(object):
    @mock.patch("vcio.vcio_report.create_purl_vulns")
    def test_single_input_purl_xlsx_output(self, mock_create_purl_vulns):
        with open(os.path.join(test_env.test_data_dir, "vcio_report/vcio_report-purl_vulns-single-input-purl.json"), "r") as f:
            mock_create_purl_vulns.return_value = json.load(f)
        input_txt = test_env.get_test_loc(
            "vcio_report/vcio_report-input-single-input-purl.txt")
        expected_result_file = test_env.get_test_loc(
            "vcio_report/vcio_report-expected-output-single-input-purl.xlsx"
        )
        actual_result_file = test_env.get_temp_file(
            "vcio_report-actual-output-single-input-purl.xlsx"
        )
        api_key = "123"
        options = [input_txt, actual_result_file, api_key]
        runner = CliRunner()
        _ = runner.invoke(vcio_report.cli, options, catch_exceptions=False)
        check_results(actual_result_file, expected_result_file, regen=False)

        wb = openpyxl.load_workbook(
            filename=actual_result_file, read_only=True)
        sheet_names = wb.sheetnames
        assert sheet_names == ['VCID_CPE', 'VULN_PACKAGES', 'VULN_FIXES']

        ws = wb['VULN_PACKAGES']
        VULN_PACKAGES_header_row_values = [cell.value for cell in ws[1]]
        assert VULN_PACKAGES_header_row_values == ['Package (PURL)', 'Present in VCIO', 'Package Type', 'VCID', 'VulnID (alias)',
                                                   'Immediate Fix', 'Non_vulnerable Fix', 'Severity', 'Scoring System', 'VulnID Origin', 'VulnID URL', 'Description']

    @mock.patch("vcio.vcio_report.create_purl_vulns")
    def test_multiple_input_purls_xlsx_output(self, mock_create_purl_vulns):
        with open(os.path.join(test_env.test_data_dir, "vcio_report/vcio_report-purl_vulns-multiple-input-purls.json"), "r") as f:
            mock_create_purl_vulns.return_value = json.load(f)
        input_txt = test_env.get_test_loc(
            "vcio_report/vcio_report-input-multiple-input-purls.txt")
        expected_result_file = test_env.get_test_loc(
            "vcio_report/vcio_report-expected-output-multiple-input-purls.xlsx"
        )
        actual_result_file = test_env.get_temp_file(
            "vcio_report-actual-output-multiple-input-purls.xlsx"
        )
        api_key = "123"
        options = [input_txt, actual_result_file, api_key]
        runner = CliRunner()
        _ = runner.invoke(vcio_report.cli, options, catch_exceptions=False)
        check_results(actual_result_file, expected_result_file, regen=False)


def check_results(actual_result_file, expected_result_file, regen=False):
    """
    Compare actual result file with expected result file.
    """
    result_wb = openpyxl.load_workbook(actual_result_file)
    if regen:
        # We overwrite `expected_result_file` with the contents of `actual_result_file`
        # for the purpose of regenerating test files
        shutil.copy2(actual_result_file, expected_result_file)
    expected_wb = openpyxl.load_workbook(expected_result_file)
    # Load named worksheets from workbooks
    results = result_wb["VULN_PACKAGES"]
    expected = expected_wb["VULN_PACKAGES"]
    # Look at each row from both worksheets at the same time
    for result_row, expected_row in zip(results.iter_rows(), expected.iter_rows()):
        # Look at each cell from the row we are looking at from above
        for result_cell, expected_cell in zip(result_row, expected_row):
            # Make sure we have the same value in each cell
            assert result_cell.value == expected_cell.value
    # compare the number of rows in the 2 sheets
    results_rows = results.max_row
    expected_rows = expected.max_row
    assert results_rows == expected_rows

    # This is another way to compare the row count but does not use the openpyxl 'max_row' method.
    results_count = 0
    for row in results:
        if not all([cell.value == None for cell in row]):
            results_count += 1
    expected_count = 0
    for row in expected:
        if not all([cell.value == None for cell in row]):
            expected_count += 1
    assert results_count == expected_count
