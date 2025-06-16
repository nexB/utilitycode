#!/usr/bin/env python
# -*- coding: utf8 -*-

# ============================================================================
#  Copyright (c) nexB Inc. http://www.nexb.com/ - All rights reserved.
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#      http://www.apache.org/licenses/LICENSE-2.0
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#
#  SPDX-License-Identifier: Apache-2.0
# ============================================================================

from collections import namedtuple

import click
import io
import os
import re
import sys

from license_expression import Licensing

from openpyxl import load_workbook

from pathlib import Path

from deja import api
from utilitycode.bom_utils import create_xlsx_output
from utilitycode.bom_utils import get_header_fields_and_index_dict

HOST_licenses = 'https://enterprise.dejacode.com/api/v2/licenses/'

# TODO: add LicenseDB as default DB -- it's currently DejaCode -- AFTER LicenseDB has been
# updated to add required `attribution` and `redistribution` fields; and update error messages
# accordingly to reflect LicenseDB as source when it's used in the LCAT process
# TODO: will also need to add option to use DejaCode when we add LicenseDB as default DB


LICENSE_HEADERS = [
    "ScanCode License Category",
    "Normalized ScanCode License Category",
    "ScanCode License Short Name",
    "Attribution",
    "Redistribution",
    "SPDX Short Identifier"
]


def add_data_to_expression(license_expression, license_data, owner_name):
    """
    Return a license expression object enriched with the given `license_data`.
    """
    # create custom Licensing Symbol to hold various DejaCode License Gallery data
    LicenseSymbol = namedtuple(
        "LicenseSymbol",
        [
            "key",
            "name",
            "is_exception",
            "category",
            "attribution",
            "redistribution",
            "spdx_short_identifier"
        ],
    )

    symbols = []
    for license_key, data in license_data.items():
        # Handle commercial/prop cases (Issue #233)
        if license_key == "commercial-license":
            name = "{} Commercial".format(owner_name)
        elif license_key == "proprietary-license":
            name = "{} Proprietary".format(owner_name)
        else:
            name = data.get("short_name", "")

        attribution = data.get("attribution_required", "")
        symbols.append(
            LicenseSymbol(
                key=license_key,
                name=name,
                is_exception=data.get("is_exception", False),
                category=data.get("category", ""),
                attribution=attribution,
                redistribution=data.get("redistribution_required", ""),
                spdx_short_identifier=data.get("spdx_license_key", "")
            )
        )

    return Licensing(symbols).parse(license_expression)


def get_cell_value(ws, col_letter, row_index):
    """
    Get the value for the specific cell
    """
    if not col_letter:
        return ''
    cell_index = col_letter + row_index
    return ws[cell_index].value


def generate_license_data(lic_exp, owner_name, license_data):
    category = ""
    lic_short_name = ""
    attribution = ""
    redistribution = ""
    spdx_short_identifier = ""

    if not lic_exp:
        return category, lic_short_name, attribution, redistribution, spdx_short_identifier
    dedup_exp = Licensing().dedup(lic_exp)
    if dedup_exp is None:
        return category, lic_short_name, attribution, redistribution, spdx_short_identifier
    dedup_exp_text = dedup_exp.render()
    try:
        updated_expression = add_data_to_expression(
            dedup_exp_text, license_data, owner_name
        )
    except:
        print(
            "WARNING: Failed to encode simplified_expression: " + dedup_exp
        )
        return category, lic_short_name, attribution, redistribution, spdx_short_identifier

    try:
        category = updated_expression.render(
            template="{symbol.wrapped.category}"
        )
        lic_short_name = updated_expression.render(
            template="{symbol.wrapped.name}"
        )
        attribution = updated_expression.render(
            template="{symbol.wrapped.attribution}"
        )
        redistribution = updated_expression.render(
            template="{symbol.wrapped.redistribution}"
        )
        spdx_short_identifier = updated_expression.render(
            template="{symbol.wrapped.spdx_short_identifier}"
        )
        # Set exception license as primary for category and attribution
        category = set_with_exception_as_primary(category)
        attribution = set_with_exception_as_primary(attribution)
    except AttributeError:
        category, lic_short_name, attribution, redistribution = "", "", "", ""

    # TODO: There is probably a better way to do this
    if "True" in attribution:
        attribution = "x"
    else:
        attribution = ""

    if "True" in redistribution:
        redistribution = "x"
    else:
        redistribution = ""
    return category, lic_short_name, attribution, redistribution, spdx_short_identifier


def process_output_bom(
    input_ws,
    expression_column,
    owner_column,
    license_category_column,
    license_name_column,
    attribution_column,
    redistribution_column,
    license_data
):
    """
    Process the data and return a data list that is ready for the
    xlsx creation
    """
    headers = [cell.value for cell in input_ws[1]]
    headers.extend(LICENSE_HEADERS)

    new_bom_data = []
    new_bom_data.append(headers)

    for row in input_ws.iter_rows(min_row=2):
        new_row_data = [cell.value for cell in row]
        current_row_index = str(row[0].row)

        lic_short_name = ""
        category = ""
        normalized_cat = ""
        attribution = ""
        redistribution = ""
        spdx_short_identifier = ""

        lic_exp = get_cell_value(
            input_ws, expression_column, current_row_index)
        owner_name = get_cell_value(input_ws, owner_column, current_row_index)
        category, lic_short_name, attribution, redistribution, spdx_short_identifier = generate_license_data(
            lic_exp, owner_name, license_data)

        # Keep all the original value if specific option is flagged
        existing_lic_short_name = get_cell_value(
            input_ws, license_name_column, current_row_index)
        existing_category = get_cell_value(
            input_ws, license_category_column, current_row_index)
        existing_attribution = get_cell_value(
            input_ws, attribution_column, current_row_index)
        existing_redistribution = get_cell_value(
            input_ws, redistribution_column, current_row_index)

        if license_name_column and existing_lic_short_name:
            lic_short_name = existing_lic_short_name
        if license_category_column and existing_category:
            category = existing_category
        if attribution_column and existing_attribution:
            attribution = existing_attribution
        if redistribution_column and existing_redistribution:
            redistribution = existing_redistribution

        if category:
            normalized_cat = deduplicate_and_simplify(category)

        new_row_data.extend(
            [category, normalized_cat, lic_short_name, attribution, redistribution, spdx_short_identifier])
        new_bom_data.append(new_row_data)
    return new_bom_data


def set_with_exception_as_primary(expression):
    """
    Return an expression string with the 'with exception' as the primary value
    For instance,
    key1 WITH key2 exception AND key 3 will become key2 exception AND key 3
    """
    keys = expression.split()
    keys_list = []
    for k in keys:
        if k.startswith("("):
            keys_list.append("(")
            keys_list.append(k.strip("("))
        elif k.endswith(")"):
            keys_list.append(k.strip(")"))
            keys_list.append(")")
        else:
            keys_list.append(k)

    updated_expression = []
    for key in keys_list:
        if key == "WITH":
            # Remove the last element from the updated_expression list to treat
            # the WITH exception as the primary
            updated_expression.pop()
        else:
            updated_expression.append(key)
    return " ".join(updated_expression).replace("( ", "(").replace(" )", ")")


def deduplicate_and_simplify(expression):
    """
    Deduplicate the expression and return the deduped expression while preserving order.
    """
    # Remove any extra spaces
    expression = re.sub(r'\s+', ' ', expression.strip())

    # Split the expression into parts
    parts = re.split(r'(\bAND\b|\bOR\b|\(|\))', expression)

    # Remove empty parts and extra spaces
    parts = [part.strip() for part in parts if part.strip()]

    # Initialize the stack for evaluation
    stack = []
    operators = []

    def process_stack():
        if len(stack) < 2:
            return
        right = stack.pop()
        left = stack.pop()
        operator = operators.pop()
        if operator == 'AND':
            # Using list to preserve order and deduplicate
            combined = list(dict.fromkeys(
                left.split(' AND ') + right.split(' AND ')))
            stack.append(' AND '.join(combined))
        elif operator == 'OR':
            left_set = set(left.split(' AND '))
            right_set = set(right.split(' AND '))
            if left_set.issubset(right_set):
                stack.append(right)
            elif right_set.issubset(left_set):
                stack.append(left)
            else:
                stack.append(f'({left} OR {right})')

    for part in parts:
        if part in {'AND', 'OR'}:
            operators.append(part)
        elif part == '(':
            operators.append(part)
        elif part == ')':
            while operators and operators[-1] != '(':
                process_stack()
            operators.pop()  # Remove the '('
        else:
            stack.append(part)
            while len(stack) > 1 and operators and operators[-1] != '(':
                process_stack()

    # Process any remaining operators
    while operators:
        process_stack()

    # The stack list should only contain one element.
    # Return the string expression (The first element)
    # instead of a list
    return stack[0] if stack else ""


def parse_expression(expression, coordinate):
    """
    Return a list of license_key values from a given expression value.
    """
    license_keys = []
    errors = ''
    if not expression:
        return [], errors
    try:
        licensing = Licensing().dedup(expression)
        if licensing is None:
            return [], errors
        for symbol in licensing.get_symbols():
            if type(symbol).__name__ == "LicenseWithExceptionSymbol":
                license_keys.append(symbol.license_symbol.key)
                license_keys.append(symbol.exception_symbol.key)
            else:
                license_keys.append(symbol.key)
    except:
        msg = "WARNING: Failed to decode license_expression: " + \
            expression + ' at: ' + coordinate
        errors = msg

    return license_keys, errors


def validate_license_expression(expression_infos, license_dict):
    """
    Check if all the licenses in the expression_infos are covered in
    the license_dict. Licenses that do not exist in the licese_dict
    are invalid and need to throw error
    """
    errors = []
    for info in expression_infos:
        # Check to see if our license key is in the set of license data we got from DejaCode
        if info.key not in license_dict:
            msg = 'Invalid license_expression value: "{}" at coord: "{}"'.format(
                info.key, info.coord
            )
            errors.append(msg)
    return errors


def collect_expression_info(ws, expression_col):
    """
    Given a worksheet and expression_col value, return a list of
    LicenseExpressionInfo namedtuples and errors if any.
    Each of these LicenseExpressionInfo objects stores the license_key
    values of a license_expression string, along with its BOM coordinate.
    """
    LicenseExpressionInfo = namedtuple(
        "LicenseExpressionInfo", ["coord", "key"])
    infos = []
    errors = []
    has_newline = False
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter != expression_col:
                continue
            if cell.value:
                # Replace the newline character
                if '\n' in cell.value:
                    cell.value = cell.value.replace('\n', '').strip()
                    has_newline = True
                lic_list, errs = parse_expression(cell.value, cell.coordinate)
                if errs:
                    errors.append(errs)
                    continue
                for key in lic_list:
                    infos.append(LicenseExpressionInfo(
                        coord=cell.coordinate, key=key))
            else:
                msg = 'Empty license_expression value at ' + cell.coordinate
                errors.append(msg)
    if has_newline:
        print("Newline character detected in the license expression field and has been replaced. Please use the generated output.")
    return infos, errors


def report_errors(errors, force, log_file_loc):
    """
    Write errors to stderr and log them to log_file_loc.
    """
    with io.open(log_file_loc, 'w', encoding='utf-8', errors='replace') as lf:
        for error in errors:
            sys.stderr.write(error + "\n")
            # logging.error(error)
            lf.write(error + '\n')
        print(("Error log: " + log_file_loc))
    if errors and not force:
        sys.stderr.write("Exiting. No output is generated.\n")
        sys.exit(1)


@click.command()
@click.option("-e", "--expression-column-name", nargs=1, required=True)
@click.option("-o", "--owner-column-name", nargs=1, required=True)
@click.option("-lc", "--license-category-column-name", nargs=1)
@click.option("-ln", "--license-name-column-name", nargs=1)
@click.option("-a", "--attribution-column-name", nargs=1)
@click.option("-r", "--redistribution-column-name", nargs=1)
@click.option(
    "-f", "--force", is_flag=True, help="Force lcat to run even if there are errors"
)
@click.option(
    '-ws', '--worksheet', nargs=1,
    help='Define the name of the worksheet to work on.'
)
@click.argument("location", type=click.Path(exists=True, readable=True))
@click.argument("destination", type=click.Path(exists=False), required=True)
@click.argument("api_key", envvar="DJE_API_KEY")
def cli(
    expression_column_name,
    owner_column_name,
    license_category_column_name,
    license_name_column_name,
    attribution_column_name,
    redistribution_column_name,
    force,
    worksheet,
    location,
    destination,
    api_key,
):
    """
    Retrieve details—including category, license name, attribution, and
    redistribution value—from DejaCode using the license expression found
    in the input.
    """
    api_key = api_key.strip("'").strip('"')

    input_bom = load_workbook(location)
    if worksheet:
        input_ws = input_bom[worksheet]
    else:
        input_ws = input_bom.active
    header_dict = get_header_fields_and_index_dict(input_ws)

    if expression_column_name not in header_dict:
        click.echo(
            "The entered expression column name field cannot be found in the input: " + expression_column_name)
        sys.exit(1)
    if owner_column_name not in header_dict:
        click.echo(
            "The entered owner column name field cannot be found in the input: " + owner_column_name)
        sys.exit(1)
    if license_category_column_name and license_category_column_name not in header_dict:
        click.echo("The entered license category column name field cannot be found in the input: " +
                   license_category_column_name)
        sys.exit(1)
    if license_name_column_name and license_name_column_name not in header_dict:
        click.echo(
            "The entered license name column name field cannot be found in the input: " + license_name_column_name)
        sys.exit(1)
    if attribution_column_name and attribution_column_name not in header_dict:
        click.echo(
            "The entered attribution column name field cannot be found in the input: " + attribution_column_name)
        sys.exit(1)
    if redistribution_column_name and redistribution_column_name not in header_dict:
        click.echo("The entered redistribution column name field cannot be found in the input: " +
                   redistribution_column_name)
        sys.exit(1)

    expression_column = header_dict[expression_column_name]
    owner_column = header_dict[owner_column_name]

    license_category_column = ''
    license_name_column = ''
    attribution_column = ''
    redistribution_column = ''
    if license_category_column_name:
        license_category_column = header_dict[license_category_column_name]
    if license_name_column_name:
        license_name_column = header_dict[license_name_column_name]
    if attribution_column_name:
        attribution_column = header_dict[attribution_column_name]
    if redistribution_column_name:
        redistribution_column = header_dict[redistribution_column_name]

    errors = []
    expression_info, err = collect_expression_info(input_ws, expression_column)

    if err:
        errors.extend(err)

    # Get a de-duped list of license_key values to query DJE
    license_keys = list(set([le.key for le in expression_info]))

    license_dict = {}

    fatal_errors = [
        "Authorization denied. Invalid '--api_key'.",  "Invalid '--api_url'."]
    api_error = False
    api_url = HOST_licenses

    for lic_key in license_keys:
        lic_data, error = api.get_licenses_dje(lic_key, api_url, api_key)
        if error:
            if error in fatal_errors:
                if error == "Invalid '--api_url'.":
                    errors.append('URL not reachable: ' + api_url)
                else:
                    errors.append("Authorization denied. Invalid '--api_key'.")
                api_error = True
                break
            errors.append(error)
        if lic_data:
            license_dict[lic_data["key"]] = lic_data

    if not api_error and expression_info:
        errors.extend(validate_license_expression(
            expression_info, license_dict))

    output_path = Path(destination)
    log_file_loc = os.path.join(output_path.parent.absolute(), 'error.log')

    if errors:
        report_errors(errors, force, log_file_loc)

    # create the new ouput bom and apply license_data
    new_bom_data = process_output_bom(
        input_ws,
        expression_column,
        owner_column,
        license_category_column,
        license_name_column,
        attribution_column,
        redistribution_column,
        license_dict
    )

    create_xlsx_output(destination, new_bom_data, input_bom, input_ws)
    sys.stdout.write('Saving output BOM to: {}\n'.format(destination))
