#!/usr/bin/env python
# -*- coding: utf8 -*-

# ============================================================================
#  Copyright (c) nexB Inc. http://www.nexb.com/ - All rights reserved.
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#      http://www.apache.org/licenses/LICENSE-2.0
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#
#  SPDX-License-Identifier: Apache-2.0
# ============================================================================

from __future__ import absolute_import
from __future__ import print_function
from __future__ import division
from __future__ import unicode_literals

from collections import Counter
import sys

import click
from openpyxl import load_workbook

from utilitycode.bom_utils import create_xlsx_output
from utilitycode.bom_utils import curate_value
from utilitycode.bom_utils import get_header_fields_and_index_dict


# silence some python3 Click warnings
click.disable_unicode_literals_warning = True


def component_rows(worksheet, counter, columns):
    """
    Yield key, value pairs where the key is the "component" tuple and the
    value is the row data for that component.
    """
    for row in worksheet.iter_rows(min_row=2):
        component = tuple(
            [curate_value(cell.value)
             for cell in row if cell.column_letter in columns]
        )
        if component in counter.keys():
            yield component, [counter.get(component)] + [
                curate_value(cell.value) for cell in row
            ]


def components(worksheet, columns):
    """
    Given a worksheet and column characters to summarize on, yield a tuple
    of data selected by the given columns characters. These tuples will act
    as the keys to a dict that will be used for summarization
    """
    for row in worksheet.iter_rows(min_row=2):
        yield tuple(
            [curate_value(cell.value)
             for cell in row if cell.column_letter in columns]
        )


@click.command()
@click.option(
    "-c",
    "--columns",
    multiple=True,
    required=True,
    help="Specify the column name you wish to summarize on.",
)
@click.argument("location", type=click.Path(exists=True, readable=True))
@click.argument("destination", type=click.Path(exists=False), required=True)
def cli(columns, location, destination):
    """
    This script creates a "Summarized Count" column, tracking the
    occurrence of values within the defined column(s).
    """
    input_bom = load_workbook(location)
    worksheet = input_bom.active
    header_dict = get_header_fields_and_index_dict(worksheet)

    columns_name = [col for col in columns]
    columns_letter = []
    for name in columns_name:
        if name not in header_dict:
            click.echo(
                "The entered name field is not found: " + name)
            sys.exit(1)
        else:
            columns_letter.append(header_dict[name])

    headers = ["Summarized Count"] + \
        [curate_value(cell.value) for cell in worksheet[1]]
    rows = {
        k: v
        for k, v in component_rows(
            worksheet, Counter(components(
                worksheet, columns_letter)), columns_letter
        )
    }

    new_bom_data = []
    new_bom_data.append(headers)
    for r in list(rows.values()):
        new_bom_data.append(r)

    create_xlsx_output(destination, new_bom_data)
    sys.stdout.write("Saving output BOM to: {}\n".format(destination))
