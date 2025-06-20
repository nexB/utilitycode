#!/usr/bin/env python
# -*- coding: utf8 -*-

# ============================================================================
#  Copyright (c) nexB Inc. http://www.nexb.com/ - All rights reserved.
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#      http://www.apache.org/licenses/LICENSE-2.0
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#
#  SPDX-License-Identifier: Apache-2.0
# ============================================================================

import json

import click
import requests
from openpyxl import Workbook
from openpyxl.styles import DEFAULT_FONT
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import borders
from openpyxl.styles.borders import Border
from packageurl import PackageURL

vcids = []


def deduplicate_purls(purls):
    """
    Deduplicate all input PURLs.  PURLs with different versions or no version
    are treated as unique for purposes of deduplication.
    """
    reviewed = set()
    deduplicated_purls = []

    for purl in purls:
        purl = purl.strip()
        if purl not in reviewed:
            reviewed.add(purl)
            deduplicated_purls.append(purl)
    return deduplicated_purls


def create_purl_vulns(purls, token):
    """
    Gather data from the '/api/packages/bulk_lookup' endpoint.
    """
    headers = {"Authorization": f"Token {token}",
               "Content-Type": "application/json"}
    request_body = {
        "purls": purls
    }
    purl_vulns_response = requests.post(
        "https://public.vulnerablecode.io/api/packages/bulk_lookup",
        data=json.dumps(request_body),
        headers=headers,
    ).json()

    purl_vulns = []
    for purl_vuln in purl_vulns_response:
        purl_vuln_map = {}
        purl_vuln_map["input_purl"] = purl_vuln.get("purl")
        purl_vuln_map["vuln_details"] = purl_vuln
        purl_vulns.append(purl_vuln_map)
    return purl_vulns


def create_xlsx(purl_vulns, destination):
    """
    Generate a .xlsx from the data gathered from the '/api/packages/bulk_lookup' endpoint.
    """
    wb_bulk_lookup = Workbook()
    wb_bulk_lookup_sheetnames = ["VCID_CPE", "VULN_PACKAGES", "VULN_FIXES"]
    for sheet_name in wb_bulk_lookup_sheetnames:
        ws = wb_bulk_lookup.create_sheet(sheet_name)
        ws.freeze_panes = "A2"
        ws.append([f"This is {ws.title}"])

    # Populate the .xlsx sheets.
    if purl_vulns:
        for result in purl_vulns:
            input_purl = result.get('input_purl')
            true_purl = PackageURL.from_string(input_purl)
            # ID next/latest non-vulns.
            non_vuln_data = {}
            non_vuln_data["input_purl"] = input_purl
            non_vuln_data["true_purl"] = true_purl

            result_purl = result.get('vuln_details').get('purl')
            non_vuln_data["result_purl"] = result_purl
            non_vuln_packageurl = PackageURL.from_string(result_purl)

            next_non_vuln_vers = result.get('vuln_details').get(
                "next_non_vulnerable_version")
            non_vuln_data["next_non_vulnerable_version"] = None
            non_vuln_data["next_non_vulnerable_package"] = None
            if next_non_vuln_vers:
                non_vuln_data["next_non_vulnerable_version"] = next_non_vuln_vers
                non_vuln_data["next_non_vulnerable_package"] = str(
                    PackageURL(
                        name=true_purl.name,
                        type=true_purl.type,
                        version=next_non_vuln_vers,
                        qualifiers=non_vuln_packageurl.qualifiers,
                        subpath=non_vuln_packageurl.subpath,
                        namespace=true_purl.namespace,
                    )
                )

            latest_non_vuln_vers = result.get('vuln_details').get(
                "latest_non_vulnerable_version")
            non_vuln_data["latest_non_vulnerable_version"] = None
            non_vuln_data["latest_non_vulnerable_package"] = None
            if latest_non_vuln_vers:
                non_vuln_data["latest_non_vulnerable_version"] = latest_non_vuln_vers
                non_vuln_data["latest_non_vulnerable_package"] = str(
                    PackageURL(
                        name=true_purl.name,
                        type=true_purl.type,
                        version=latest_non_vuln_vers,
                        qualifiers=non_vuln_packageurl.qualifiers,
                        subpath=non_vuln_packageurl.subpath,
                        namespace=true_purl.namespace,
                    )
                )
            for ws in wb_bulk_lookup:
                if ws.title == "VCID_CPE":
                    cpe_row = 1
                    cpe_col = 1
                    header = ["VCID", "CPE"]
                    for h in header:
                        ws.cell(row=cpe_row, column=cpe_col).value = h
                        cpe_col += 1
                    for vuln in result.get('vuln_details').get("affected_by_vulnerabilities"):
                        vuln_id = vuln.get("vulnerability_id")
                        # Check whether the vuln_id -- i.e., the VCID -- is in the list 'vcids'.
                        # If not, add it to that list and add the CPE data to the VCID_CPE sheet.
                        if vuln_id not in vcids:
                            vcids.append(vuln_id)
                            for ref in vuln.get("references"):
                                if ref.get("reference_id").startswith("cpe:"):
                                    newRowLocation = ws.max_row + 1
                                    ws.cell(
                                        column=1,
                                        row=newRowLocation,
                                        value=vuln_id,
                                    )
                                    ws.cell(
                                        column=2,
                                        row=newRowLocation,
                                        value=ref.get("reference_id"),
                                    )
                elif ws.title == "VULN_PACKAGES":
                    count = 1
                    col = 1
                    header = [
                        "Package (PURL)",
                        "Present in VCIO",
                        "Package Type",
                        "VCID",
                        "VulnID (alias)",
                        "Immediate Fix",
                        "Non_vulnerable Fix",
                        "Severity",
                        "Scoring System",
                        "VulnID Origin",
                        "VulnID URL",
                        "Description",
                    ]
                    for h in header:
                        ws.cell(row=count, column=col).value = h
                        col += 1
                    purl_id = result.get('vuln_details').get("purl")
                    # Get the non-vuln Packages for this PURL -- 1 time only for each PURL,
                    # not each vuln, and only if the PURL has 1+ affected by vulnerabilities.
                    if result.get('vuln_details').get("affected_by_vulnerabilities"):
                        # Identify the Packages for the 'Immediate Fix' column.
                        for vuln in result.get('vuln_details').get("affected_by_vulnerabilities"):
                            vuln_fixed_packages = vuln["fixed_packages"]
                            immediate_fixes = []
                            non_vuln_immediate_fixes = []
                            next_and_latest_non_vuln_fixes = []
                            if non_vuln_data.get("next_non_vulnerable_package"):
                                next_and_latest_non_vuln_fixes.append(
                                    non_vuln_data.get("next_non_vulnerable_package"))
                            if non_vuln_data.get("latest_non_vulnerable_package"):
                                next_and_latest_non_vuln_fixes.append(
                                    non_vuln_data.get("latest_non_vulnerable_package"))
                            vuln_fixes = {}
                            vuln_fixes["purl"] = true_purl
                            for fix in vuln_fixed_packages:
                                immediate_fixes.append(fix.get("purl"))
                                if fix.get("is_vulnerable") is not True:
                                    non_vuln_immediate_fixes.append(
                                        fix.get("purl"))

                            deduped_next_and_latest_non_vuln_fixes = deduplicate_purls(
                                next_and_latest_non_vuln_fixes)
                            for ref in vuln.get("references"):
                                newRowLocation = ws.max_row + 1
                                ws.cell(
                                    column=1,
                                    row=newRowLocation,
                                    value=purl_id,
                                )
                                # column 2 = 'Present in VCIO'
                                package_id = result.get('vuln_details').get(
                                    "url").split("/")[-1]
                                ws.cell(
                                    column=2,
                                    row=newRowLocation,
                                    value=package_id,
                                )
                                # column 3 = 'Package Type'
                                package_type = true_purl.type
                                ws.cell(
                                    column=3,
                                    row=newRowLocation,
                                    value=package_type,
                                )
                                # column 4 = 'VCID'
                                vuln_id = vuln["vulnerability_id"]
                                ws.cell(
                                    column=4,
                                    row=newRowLocation,
                                    value=vuln_id,
                                )
                                # column 5 = 'VulnID (alias)'
                                vuln_alias = ""
                                if (
                                    ref.get("url").startswith(
                                        "https://access.redhat.com/errata/"
                                    )
                                    or ref.get("url").startswith(
                                        "https://access.redhat.com/security/cve/"
                                    )
                                    or ref.get("url").startswith(
                                        "https://github.com/advisories"
                                    )
                                    or ref.get("url").startswith(
                                        "https://nvd.nist.gov/vuln/detail"
                                    )
                                ):
                                    vuln_alias = ref.get("url").split("/")[-1]
                                elif ref.get("url").startswith(
                                    "https://bugzilla.redhat.com/"
                                ):
                                    vuln_alias = ref.get("url").split(
                                        "show_bug.cgi?id="
                                    )[-1]
                                elif ref.get("url").startswith(
                                    "https://bugs.debian.org/cgi-bin/"
                                ):
                                    vuln_alias = ref.get("url").split(
                                        "bugreport.cgi?bug="
                                    )[-1]
                                elif ref.get("url").startswith(
                                    "https://cve.mitre.org/cgi-bin/"
                                ):
                                    vuln_alias = ref.get("url").split(
                                        "cvename.cgi?name="
                                    )[-1]
                                elif ref.get("url").startswith(
                                    "https://access.redhat.com/hydra/rest/securitydata/cve/"
                                ) and ref.get("url").endswith(".json"):
                                    vuln_alias = (
                                        ref.get("url").split(
                                            "/")[-1].rstrip(".json")
                                    )
                                elif ref.get("url").startswith(
                                    "http://people.canonical.com/~ubuntu-security/cve/"
                                ) and ref.get("url").endswith(".html"):
                                    vuln_alias = (
                                        ref.get("url").split(
                                            "/")[-1].rstrip(".html")
                                    )
                                elif ref.get("url").startswith(
                                    "http://activemq.apache.org/security-advisories.data/"
                                ) and ref.get("url").endswith("-announcement.txt"):
                                    vuln_alias = (
                                        ref.get("url")
                                        .split("/")[-1]
                                        .rstrip("-announcement.txt")
                                    )
                                else:
                                    vuln_alias = ""
                                # column 5 = 'VulnID (alias)'
                                ws.cell(
                                    column=5,
                                    row=newRowLocation,
                                    value=vuln_alias,
                                )
                                # column 6 = "Immediate Fix"
                                ws.cell(
                                    column=6,
                                    row=newRowLocation,
                                    value="\n".join(immediate_fixes),
                                )
                                # column 7 = "Non_vulnerable Fix"
                                ws.cell(
                                    column=7,
                                    row=newRowLocation,
                                    value="\n".join(
                                        deduped_next_and_latest_non_vuln_fixes
                                    ),
                                )
                                # column 8 = "Severity" and column 9 = "Scoring System"
                                vuln_severity = ""
                                ref_scores = ref.get("scores")
                                if ref_scores:
                                    for ref_score in ref_scores:
                                        if (
                                            ref_score.get("scoring_system") == "archlinux"
                                            or ref_score.get("scoring_system") == "cvssv2"
                                            or ref_score.get("scoring_system") == "cvssv3"
                                            or ref_score.get("scoring_system") == "cvssv3.1"
                                            or ref_score.get("scoring_system") == "cvssv3.1_qr"
                                            or ref_score.get("scoring_system") == "cvssv4"
                                            or ref_score.get("scoring_system") == "epss"
                                            or ref_score.get("scoring_system") == "generic_textual"
                                            or ref_score.get("scoring_system") == "rhas"
                                            or ref_score.get("scoring_system") == "rhbs"
                                            or ref_score.get("scoring_system") == "ssvc"
                                        ):
                                            vuln_severity = ref_score.get(
                                                "value")
                                            vuln_scoring_system = ref_score.get(
                                                "scoring_system"
                                            )
                                        else:
                                            vuln_severity = "TBD"
                                            vuln_scoring_system = "TBD"
                                else:
                                    vuln_severity = "NA"
                                    vuln_scoring_system = "NA"
                                # column 8 = "Severity"
                                ws.cell(
                                    column=8,
                                    row=newRowLocation,
                                    value=vuln_severity,
                                )
                                # column 9 = "Scoring System"
                                ws.cell(
                                    column=9,
                                    row=newRowLocation,
                                    value=vuln_scoring_system,
                                )
                                # column 10 = 'VulnID Origin' (was column 9)
                                vuln_origin = ""
                                if vuln_alias.startswith("CVE"):
                                    vuln_origin = "NVD"
                                elif vuln_alias.startswith("GHSA"):
                                    vuln_origin = "GHSA"
                                else:
                                    vuln_origin = "Other"
                                ws.cell(
                                    column=10,
                                    row=newRowLocation,
                                    value=vuln_origin,
                                )
                                # column 11 = 'VulnID URL' (was column 10)
                                ws.cell(
                                    column=11,
                                    row=newRowLocation,
                                    value=ref.get("reference_url"),
                                )
                                # column 12 = 'Description' (was column 11)
                                vuln_summary = vuln["summary"]
                                ws.cell(
                                    column=12,
                                    row=newRowLocation,
                                    value=vuln_summary,
                                )
                elif ws.title == "VULN_FIXES":
                    count = 1
                    col = 1
                    header = [
                        "Package (PURL)",
                        "Package Type",
                        "VCID",
                        "Immediate Fix",
                        "Non_vulnerable Fix",
                        "Description",
                    ]
                    for h in header:
                        ws.cell(row=count, column=col).value = h
                        col += 1

                    if result.get("vuln_details").get("affected_by_vulnerabilities"):
                        # Identify the Packages for the 'Immediate Fix' column.
                        for vuln in result.get("vuln_details").get("affected_by_vulnerabilities"):
                            vuln_fixed_packages = vuln["fixed_packages"]
                            immediate_fixes = []
                            for fix in vuln_fixed_packages:
                                immediate_fixes.append(fix.get("purl"))

                            newRowLocation = ws.max_row + 1
                            # column 1 = 'Package (PURL)'
                            ws.cell(
                                column=1,
                                row=newRowLocation,
                                value=purl_id,
                            )
                            # column 2 = 'Package Type'
                            package_type = true_purl.type
                            ws.cell(
                                column=2,
                                row=newRowLocation,
                                value=package_type,
                            )
                            # column 3 = 'VCID'
                            vuln_id = vuln["vulnerability_id"]
                            ws.cell(
                                column=3,
                                row=newRowLocation,
                                value=vuln_id,
                            )
                            # column 4 = "Immediate Fix"
                            ws.cell(
                                column=4,
                                row=newRowLocation,
                                value="\n".join(immediate_fixes),
                            )
                            # column 5 = "Non_vulnerable Fix"
                            ws.cell(
                                column=5,
                                row=newRowLocation,
                                value="\n".join(
                                    deduped_next_and_latest_non_vuln_fixes
                                ),
                            )
                            # column 6 = 'Description'
                            vuln_summary = vuln["summary"]
                            ws.cell(
                                column=6,
                                row=newRowLocation,
                                value=vuln_summary,
                            )

    sh = wb_bulk_lookup["Sheet"]
    wb_bulk_lookup.remove(sh)
    # We handle the header row in more detail below using cell.font = font.
    DEFAULT_FONT.size = 10
    for ws in wb_bulk_lookup.worksheets:
        # Get the maximum number of columns in the worksheet.
        max_columns = ws.max_column
        # Iterate over the columns and set a default width.
        for column in range(1, max_columns + 1):
            ws.column_dimensions[ws.cell(
                row=1, column=column).column_letter].width = 25
        # Set individual column widths as needed.
        if ws.title == "VCID_CPE":
            ws.column_dimensions["B"].width = 42
        if ws.title == "VULN_PACKAGES":
            ws.column_dimensions["A"].width = 42
            ws.column_dimensions["B"].width = 10
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 20
            ws.column_dimensions["E"].width = 20
            ws.column_dimensions["F"].width = 50
            ws.column_dimensions["G"].width = 50
            ws.column_dimensions["H"].width = 12
            ws.column_dimensions["I"].width = 15
            ws.column_dimensions["J"].width = 12
            ws.column_dimensions["K"].width = 60
            ws.column_dimensions["L"].width = 75
        if ws.title == "VULN_FIXES":
            ws.column_dimensions["A"].width = 42
            ws.column_dimensions["B"].width = 10
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 50
            ws.column_dimensions["E"].width = 50
            ws.column_dimensions["F"].width = 75
        font = Font(name="Calibri", size=10, bold=True, italic=False)
        for rows in ws.iter_rows(min_row=1, max_row=1, min_col=None):
            for cell in rows:
                cell.fill = PatternFill(
                    start_color="CCE6FF",
                    end_color="CCE6FF",
                    fill_type="solid",
                )
                cell.font = font
        border = borders.Side(
            style=None, color="FF000000", border_style="thin")
        thin = Border(left=border, right=border, bottom=border, top=border)
        alignment = Alignment(
            wrapText=True, horizontal="left", vertical="center")
        for row in ws.iter_rows(min_row=None, min_col=None, max_row=None, max_col=None):
            for cell in row:
                cell.border = thin
                cell.alignment = alignment

    wb_bulk_lookup.save(destination)


@click.command()
@click.argument("location", type=click.Path(exists=True, readable=True))
@click.argument("destination", type=click.Path(exists=False), required=True)
@click.argument("api_key", envvar="VCIO_API_KEY")
@click.help_option('-h', '--help')
def cli(
    location,
    destination,
    api_key,
):
    """
    vcio_report is a command-line utility that enables a user to submit a
    .txt containing a list of one or more syntactically-correct PURLs and
    retrieve a .xlsx of vulnerabilities (if any) for each listed PURL
    contained in the VCIO database.
    """
    with open(location) as f:
        purls = f.read().splitlines()
    all_purl_vulns = create_purl_vulns(purls, api_key)
    create_xlsx(all_purl_vulns, destination)
