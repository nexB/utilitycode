#!/usr/bin/env python
# -*- coding: utf8 -*-

# ============================================================================
#  Copyright (c) nexB Inc. http://www.nexb.com/ - All rights reserved.
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#      http://www.apache.org/licenses/LICENSE-2.0
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#
#  SPDX-License-Identifier: Apache-2.0
# ============================================================================

import click
import sys

from copy import copy
from urllib.parse import urlparse
from urllib.request import urlopen

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from deja import api
from reporting.license_category import collect_expression_info
from utilitycode.bom_utils import get_header_fields_and_index_dict
from utilitycode.bom_utils import get_lic_key_and_coord
from utilitycode.bom_utils import get_value_and_coord
from utilitycode.utils import validate_url


def get_problematic_lic(license_data, lic_key_coord_dict):
    """
    Return a dictionary of the problematic license keys
    The dictionary key is the value of the lic_key, the dictionary value is
    the coordinate location of the key.
    """
    problematic_lic_dict = {}
    for lic_key in lic_key_coord_dict:
        if lic_key not in license_data:
            problematic_lic_dict[lic_key] = lic_key_coord_dict[lic_key]
    return problematic_lic_dict


def write_output_xlsx(header_dict, problematic_lic_dict, problematic_urls_dict,
                      redirected_urls_dict, input_ws, destination):
    """
    Write the output
    """
    output_bom = Workbook()
    output_ws = output_bom.active

    max_row = input_ws.max_row
    max_col = input_ws.max_column

    # Create header fields
    header_fields = ['Attention']
    for header in header_dict:
        header_fields.append(header)

    # Write header fields
    for index, header in enumerate(header_fields):
        output_ws.cell(row=1, column=index + 1).value = header

    # Write the problematic license key in the "Attention" field
    for invalid_lic in problematic_lic_dict:
        for row in problematic_lic_dict[invalid_lic]:
            msg = 'Invalid license: ' + invalid_lic
            row_index = int(''.join(filter(str.isdigit, row)))
            output_ws.cell(row=row_index, column=1).value = msg

    # Write the problematic url in the "Attention" field
    for invalid_url in problematic_urls_dict:
        for row in problematic_urls_dict[invalid_url]:
            # Get the index coordinate
            row_index = int(''.join(filter(str.isdigit, row)))
            # Get the letter coordinate and retrieve the field name
            col_letter = ''.join(filter(str.isalpha, list(row)))
            url_field_name = ''
            for field in header_dict:
                if col_letter == header_dict[field]:
                    url_field_name = field
                    break

            msg = 'Invalid ' + url_field_name + ': ' + invalid_url
            if not output_ws.cell(row=row_index, column=1).value:
                output_ws.cell(row=row_index, column=1).value = msg
            else:
                output_ws.cell(row=row_index, column=1).value = output_ws.cell(
                    row=row_index, column=1).value + '\n' + msg

            # Apply text wrapping
            output_ws.cell(row=row_index, column=1).alignment = Alignment(
                wrap_text=True)

    # Write the redirect url in the "Attention" field
    for redirect_url in redirected_urls_dict:
        redirect_url_value = redirected_urls_dict[redirect_url][0]

        for row in redirected_urls_dict[redirect_url][1]:
            # Get the index coordinate
            row_index = int(''.join(filter(str.isdigit, row)))
            # Get the letter coordinate and retrieve the field name
            col_letter = ''.join(filter(str.isalpha, list(row)))
            url_field_name = ''
            for field in header_dict:
                if col_letter == header_dict[field]:
                    url_field_name = field
                    break

            msg = url_field_name + ' is redirected to: ' + redirect_url_value
            if not output_ws.cell(row=row_index, column=1).value:
                output_ws.cell(row=row_index, column=1).value = msg
            else:
                output_ws.cell(row=row_index, column=1).value = output_ws.cell(
                    row=row_index, column=1).value + '\n' + msg

            # Apply text wrapping
            output_ws.cell(row=row_index, column=1).alignment = Alignment(
                wrap_text=True)

    # Set the attention column width to '60'
    attention_col_width = 60
    output_ws.column_dimensions['A'].width = attention_col_width

    # Copy the content from input to destination
    for i in range(1, max_row + 1):
        for j in range(1, max_col + 1):
            # reading cell value from source excel file
            cell = input_ws.cell(row=i, column=j)

            # write value
            if i == 1:
                new_cell = output_ws.cell(row=1, column=j + 1)
            else:
                if cell.value:
                    new_value = str(cell.value).strip()
                else:
                    new_value = cell.value
                new_cell = output_ws.cell(
                    row=i, column=j + 1, value=new_value)

            # Write formatting
            output_ws.column_dimensions[get_column_letter(j)].auto_size = True
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    click.echo('Saving BOM to %s' % destination)

    # saving the destination excel file
    output_bom.save(destination)


@click.command()
@click.argument('location', type=click.Path(exists=True, readable=True))
@click.argument('destination', type=click.Path(exists=False), required=True)
@click.option('--api_url',
              nargs=1,
              type=click.STRING,
              metavar='URL',
              help='URL to DejaCode License Library.')
@click.option('--api_key',
              nargs=1,
              type=click.STRING,
              metavar='KEY',
              help='API Key for the DejaCode License Library')
@click.option('-e', '--expression-field-name', nargs=1, metavar='FIELD_NAME',
              help='Name of the license expression field')
@click.option('-url', '--url-field-name', nargs=1, multiple=True, metavar='FIELD_NAME',
              help='Name of the URL field')
@click.help_option('-h', '--help')
def cli(api_url, api_key, expression_field_name, url_field_name, location, destination):
    """
    This utility validate the license expression field and url fields from
    the XLSX input and return the same XLSX file with the Attention column.

    Note: license expression can be validated with the choice from DejaCode
    License Library, or ScanCode LicenseDB (default)
    """
    # Check if both api_url and api_key present
    if api_url or api_key:
        if not api_url:
            msg = '"--api_url" is required.'
            click.echo(msg)
            sys.exit(1)
        if not api_key:
            msg = '"--api_key" is required.'
            click.echo(msg)
            sys.exit(1)
    else:
        api_url = ''
        api_key = ''
    api_url = api_url.strip("'").strip('"')
    api_key = api_key.strip("'").strip('"')

    input_bom = load_workbook(location)
    input_ws = input_bom.active
    header_dict = get_header_fields_and_index_dict(input_ws)
    problematic_lic_dict = {}
    problematic_urls = {}
    redirected_urls = {}

    # Check if the entered option fields exist in the input Excel file
    if url_field_name:
        not_present_field = []
        for field_name in url_field_name:
            if field_name not in header_dict:
                not_present_field.append(field_name)
        if not_present_field:
            fields = ', '.join(not_present_field)
            click.echo(
                "The url field name cannot be found in the input: " + fields)
            sys.exit(1)

    if expression_field_name:
        if expression_field_name not in header_dict:
            click.echo(
                "The expression field name cannot be found in the input: " + expression_field_name)
            sys.exit(1)

    if url_field_name:
        url_fields_dict = {}
        for field_name in url_field_name:
            url_dict = get_value_and_coord(input_bom, header_dict[field_name])
            for url in url_dict:
                if url not in url_fields_dict:
                    url_fields_dict[url] = url_dict[url]
                else:
                    for url_index in url_dict[url]:
                        url_fields_dict[url].append(url_index)
        for url in url_fields_dict:
            print("Validating: " + url)
            if not validate_url(url):
                problematic_urls[url] = url_fields_dict[url]

    if expression_field_name:
        expression_info, errors = collect_expression_info(
            input_ws, header_dict[expression_field_name])

        lic_exp_err_dict = {}
        for coord, msg in errors:
            if msg in lic_exp_err_dict:
                lic_exp_err_dict[msg].append(coord)
            else:
                lic_exp_err_dict[msg] = [coord]

        # Get a de-duped list of license_key values to query DJE
        license_keys = list(set([le.key for le in expression_info]))

        # Get the license_key and coordinate
        lic_key_coord_dict = get_lic_key_and_coord(expression_info)

        # query DejaCode License Library or ScanCode LicenseDB API for
        # license data using license_keys
        license_data = {}
        if api_url:
            fatal_errors = [
                "Authorization denied. Invalid '--api_key'.",  "Invalid '--api_url'."]
            for lic_key in license_keys:
                lic_data, error = api.get_licenses_dje(
                    lic_key, api_url, api_key)
                if error:
                    if error in fatal_errors:
                        click.echo(error)
                        sys.exit(1)
                if lic_data:
                    license_data[lic_data['key']] = lic_data
        else:
            result, _err = api.get_licenses_licensedb(license_keys)
            for data in result:
                license_data[data['key']] = data

        problematic_lic_dict = get_problematic_lic(
            license_data, lic_key_coord_dict)

        # Add the 2 dictionaries together
        problematic_lic_dict.update(lic_exp_err_dict)

    # Write to the output
    write_output_xlsx(header_dict, problematic_lic_dict,
                      problematic_urls, redirected_urls, input_ws, destination)
