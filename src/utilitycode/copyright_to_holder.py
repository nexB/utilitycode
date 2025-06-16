#!/usr/bin/env python
# -*- coding: utf8 -*-

# ============================================================================
#  Copyright (c) nexB Inc. http://www.nexb.com/ - All rights reserved.
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#      http://www.apache.org/licenses/LICENSE-2.0
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#
#  SPDX-License-Identifier: Apache-2.0
# ============================================================================

import sys

import click

from openpyxl import load_workbook

from cluecode.copyrights import CopyrightDetector
from commoncode import text
from utilitycode.bom_utils import get_header_fields_and_index_dict
from utilitycode.bom_utils import create_xlsx_output


def process_output(input_bom, copyright_column_letter, header_dict):
    """
    Create and save a new workbook containing the input_bom data with
    copyright_holder applied to each BOM row.
    """
    headers = []
    for header in header_dict:
        headers.append(header)
    headers.append('Copyright Holder')

    new_bom_data = []
    new_bom_data.append(headers)

    for row in input_bom.active.iter_rows(min_row=2):
        new_row_data = [cell.value for cell in row]
        for cell in row:
            if cell.column_letter == copyright_column_letter:
                if cell.value:
                    copyrights = text.toascii(cell.value)
                    copyright_holders = convert_copyright_to_holder(copyrights)
                    new_row_data.extend([copyright_holders])
        new_bom_data.append(new_row_data)

    return new_bom_data


def convert_copyright_to_holder(copyrights):
    """
    Given a copyright value, return the copyright holder into a single
    line, separated by comma
    """
    copyright_holders = []
    # The "CopyrightDetector().detect" doesn't get the copyright
    # holder if the statement doesn't have the "Copyright (c)". This is
    # forgivable as if the string doesn't have the "Copyright (c)", how do the
    # tool knows the string is a copyright statement.
    copyright_list = copyrights.split('\n')
    for copyright_statement in copyright_list:
        if not copyright_statement.lower().startswith('copyright (c)'):
            copyright_statement = 'copyright (c) ' + copyright_statement
        numbered_lines = [(0, copyright_statement)]
        for detection_object in CopyrightDetector().detect(numbered_lines,
                                                           include_copyrights=False, include_holders=True, include_authors=False,
                                                           include_copyright_years=False, include_copyright_allrights=False):
            holder = detection_object.holder
            # Avoid duplication
            if holder not in copyright_holders:
                copyright_holders.append(holder)
    return ', '.join(copyright_holders)


@click.command()
@click.option('-c', '--copyright-column-name', nargs=1, required=True)
@click.argument('location', type=click.Path(exists=True, readable=True))
@click.argument('destination', type=click.Path(exists=False), required=True)
@click.help_option('-h', '--help')
def cli(copyright_column_name, location, destination):
    """
    The utility generate a "Copyright Holder" field based on the
    copyright column defined by the user in the input.
    """
    input_bom = load_workbook(location)
    input_ws = input_bom.active
    header_dict = get_header_fields_and_index_dict(input_ws)

    if copyright_column_name not in header_dict:
        click.echo(
            "The entered copyright name field cannot be found in the input: " + copyright_column_name)
        sys.exit(1)

    copyright_column_letter = header_dict[copyright_column_name]

    sys.stdout.write('Saving output BOM to: {}\n'.format(destination))
    output_data = process_output(
        input_bom, copyright_column_letter, header_dict)
    create_xlsx_output(destination, output_data)
