#!/usr/bin/env python
# -*- coding: utf8 -*-

# ============================================================================
#  Copyright (c) nexB Inc. http://www.nexb.com/ - All rights reserved.
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#      http://www.apache.org/licenses/LICENSE-2.0
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#
#  SPDX-License-Identifier: Apache-2.0
# ============================================================================

from __future__ import print_function, absolute_import

from collections import Counter
from copy import copy
import csv
import json
import logging
import operator
import os
import sys
import warnings

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

logger = logging.getLogger('utilitycode')
# logging.basicConfig(level=logging.DEBUG)

# disable annoying warnings from OpenPyxl
warnings.filterwarnings('ignore', module='openpyxl')


def get_expressions(location, column_char):
    """
    Given a location path of an Excel XLSX BOM and a license expression
    column character, return list of expression strings.
    """
    workbook = openpyxl.load_workbook(location)
    if has_macros(workbook):
        print("Workbook contains macros, please remove them")
        sys.exit()

    worksheet = workbook.active

    expressions = []
    for cell in get_column(worksheet, column_char)[1:]:
        if cell.value:
            expressions.append(cell.value)

    return expressions


def create_nexb_bom(data):
    '''
    Given a destination path and data from the input csv, write out the
    report XLSX file that contains the analysis data formatted to nexb
    report BOM standards.
    '''
    side = openpyxl.styles.borders.Side(border_style='thin', color='000000')
    border = openpyxl.styles.borders.Border(
        top=side, bottom=side, left=side, right=side)

    header_font = openpyxl.styles.fonts.Font(name='Arial', size=10, b=True)
    header_color = openpyxl.styles.fills.PatternFill('solid', fgColor='c5ffff')

    content_font = openpyxl.styles.fonts.Font(name='Arial', size=10)
    content_alignment = openpyxl.styles.alignment.Alignment(
        vertical='center', wrapText=True)

    wb = openpyxl.Workbook()
    ws = wb.active

    for row in data:
        xlsx_row = []
        for cell in row:
            xlsx_cell = openpyxl.cell.cell.Cell(worksheet=ws, value=cell)
            xlsx_cell.border = border
            if row == data[0]:
                xlsx_cell.font = header_font
                xlsx_cell.fill = header_color
            else:
                xlsx_cell.font = content_font
                xlsx_cell.alignment = content_alignment
            xlsx_row.append(xlsx_cell)
        ws.append(xlsx_row)

    return wb


def create_nexb_bom_from_scancode(data, report=False):
    '''
    Given a destination path and data from the input csv, write out the
    report XLSX file that contains the analysis data formatted to nexb
    report BOM standards.
    '''
    side = openpyxl.styles.borders.Side(border_style='thin', color='000000')
    border = openpyxl.styles.borders.Border(
        top=side, bottom=side, left=side, right=side)

    header_font = openpyxl.styles.fonts.Font(name='Calibri', size=10, b=True)
    bom_font = openpyxl.styles.fonts.Font(name='Calibri', size=10)

    file_cat_color = openpyxl.styles.fills.PatternFill(
        'solid', fgColor='FFF2CC')
    status_color = openpyxl.styles.fills.PatternFill('solid', fgColor='F2F2F2')
    report_color = openpyxl.styles.fills.PatternFill('solid', fgColor='C0FEFD')
    license_color = openpyxl.styles.fills.PatternFill(
        'solid', fgColor='FFFF00')

    content_alignment = openpyxl.styles.alignment.Alignment(
        vertical='center', wrapText=True)
    vertical_alignment = Alignment(textRotation=90, horizontal="center")
    center_alignment = Alignment(horizontal="center", vertical='center')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "INVENTORY"

    for row in data:
        xlsx_row = []
        for idx, cell in enumerate(row):
            # The index is used to determine the color theme for the column
            try:
                xlsx_cell = openpyxl.cell.cell.Cell(worksheet=ws, value=cell)
            except:
                # ToDo: temp fix
                # Leave blank if the cell has encoding issue or otherwise
                # openpyxml will throw AttributionError
                xlsx_cell = openpyxl.cell.cell.Cell(worksheet=ws, value='')
            xlsx_cell.border = border
            if row == data[0]:
                xlsx_cell.font = header_font
                if report:
                    if idx <= 2:
                        xlsx_cell.fill = file_cat_color
                    elif idx >= 3 and idx <= 6:
                        xlsx_cell.fill = status_color
                    elif (idx >= 7 and idx <= 20) or (idx >= 25 and idx <= 39):
                        xlsx_cell.fill = report_color
                    elif idx >= 21 and idx <= 24:
                        xlsx_cell.fill = license_color
            else:
                xlsx_cell.font = bom_font
                xlsx_cell.alignment = content_alignment
            xlsx_row.append(xlsx_cell)
        ws.append(xlsx_row)

    if report:
        # Header formatting
        header_height = 82.00
        short_width = 4.00
        normal_width = 15.00
        long_width = 50.00

        # Set the height of the first row
        ws.row_dimensions[1].height = header_height
        index = 1
        # Columns letter
        short_width_cols = ['A', 'D', 'O', 'X',
                            'Y', 'AB', 'AC', 'AD', 'AE', 'AF']
        long_width_cols = ['L', 'AG', 'AH', 'AI', 'AJ']
        for idx, row in enumerate(ws.rows):
            for cell in row:
                if get_column_letter(cell.col_idx) in short_width_cols:
                    if idx == 0:
                        cell.alignment = vertical_alignment
                    else:
                        cell.alignment = center_alignment

        for column_cells in ws.columns:
            col_letter = get_column_letter(index)
            if col_letter in short_width_cols:
                ws.column_dimensions[col_letter].width = short_width
            elif col_letter in long_width_cols:
                ws.column_dimensions[col_letter].width = long_width
            else:
                ws.column_dimensions[col_letter].width = normal_width
            index = index + 1

    return wb


def create_scio2inv_bom(data, packages_ws, resources_ws, reorder):
    '''
    The data is a dictionary that the key is the sheetname,
    and the value is the sheet content with ready to write to xlsx formatted.
    '''
    side = openpyxl.styles.borders.Side(border_style='thin', color='000000')
    border = openpyxl.styles.borders.Border(
        top=side, bottom=side, left=side, right=side)

    header_font = openpyxl.styles.fonts.Font(
        name='Calibri', size=10, bold=True)
    bom_font = openpyxl.styles.fonts.Font(name='Calibri', size=10)
    detection_font = openpyxl.styles.fonts.Font(
        name='Calibri', size=10, italic=True)

    filecat_color = openpyxl.styles.fills.PatternFill(
        'solid', fgColor='FFF2CC')
    report_color = openpyxl.styles.fills.PatternFill('solid', fgColor='C0FEFD')
    status_color = openpyxl.styles.fills.PatternFill('solid', fgColor='F2F2F2')

    content_alignment = openpyxl.styles.alignment.Alignment(
        vertical='center', wrapText=True)
    vertical_alignment = Alignment(textRotation=90, horizontal="center")
    center_alignment = Alignment(horizontal="center", vertical='center')

    # Header formatting
    header_height = 82.00
    short_width = 4.00
    normal_width = 15.00
    long_width = 50.00

    worksheets_name_list = list(data.keys())
    wb = openpyxl.Workbook()
    if not reorder:
        changelog_ws = wb.active
        changelog_ws.title = "CHANGELOG"
        changelog_ws.freeze_panes = "A2"

        changelog_ws['A1'] = 'Version'
        changelog_ws['B1'] = 'By'
        changelog_ws['C1'] = 'Changes'

        changelog_ws['A1'].font = Font(size=10, bold=True)
        changelog_ws['B1'].font = Font(size=10, bold=True)
        changelog_ws['C1'].font = Font(size=10, bold=True)
        readme_index = 1
        for dec in changelog_ws.columns:
            col_letter = get_column_letter(readme_index)
            if col_letter == 'A':
                changelog_ws.column_dimensions[col_letter].width = normal_width
            elif col_letter == 'B':
                changelog_ws.column_dimensions[col_letter].width = normal_width
            elif col_letter == 'C':
                changelog_ws.column_dimensions[col_letter].width = long_width
            readme_index = readme_index + 1

    worksheets = [wb.create_sheet(title=name) for name in worksheets_name_list]
    for i, worksheet_name in enumerate(data):
        for row in data[worksheet_name]:
            xlsx_row = []
            for idx, cell in enumerate(row):
                # The index is used to determine the color theme for the column
                try:
                    xlsx_cell = openpyxl.cell.cell.Cell(
                        worksheet=worksheets[i], value=cell)
                except:
                    # ToDo: temp fix
                    # Leave blank if the cell has encoding issue or otherwise
                    # openpyxml will throw AttributionError
                    xlsx_cell = openpyxl.cell.cell.Cell(
                        worksheet=worksheets[i], value='')
                xlsx_cell.border = border
                if row == data[worksheet_name][0]:
                    xlsx_cell.font = header_font
                    # Freeze the header row
                    worksheets[i].freeze_panes = "A2"
                    if worksheet_name == packages_ws:
                        if not reorder:
                            if idx <= 2:
                                xlsx_cell.fill = status_color
                            elif idx >= 3 and idx <= 25:
                                xlsx_cell.fill = report_color
                            if idx >= 18 and idx <= 24:
                                xlsx_cell.font = detection_font
                        else:
                            xlsx_cell.fill = report_color
                    elif worksheet_name == resources_ws:
                        if not reorder:
                            if idx <= 2:
                                xlsx_cell.fill = filecat_color
                            elif idx >= 3 and idx <= 5:
                                xlsx_cell.fill = status_color
                            elif idx >= 6 and idx <= 29:
                                xlsx_cell.fill = report_color
                            if idx >= 22 and idx <= 29:
                                xlsx_cell.font = detection_font
                        else:
                            xlsx_cell.fill = report_color
                else:
                    xlsx_cell.font = bom_font
                    xlsx_cell.alignment = content_alignment

                xlsx_row.append(xlsx_cell)
            worksheets[i].append(xlsx_row)
        # This is the fomatting for the worksheet
        # Set the height of the first row
        worksheets[i].row_dimensions[1].height = header_height
        index = 1
        if not reorder:
            if worksheet_name == packages_ws or worksheet_name == resources_ws:
                if worksheet_name == packages_ws:
                    # Columns letter for the PACKAGES sheet
                    short_width_cols = ['D', 'L', 'M', 'N']
                    long_width_cols = ['F', 'G', 'J',
                                       'K', 'O', 'P', 'W', 'X', 'Y']
                elif worksheet_name == resources_ws:
                    # Columns letter for the RESOURCES sheet
                    short_width_cols = ['A', 'G', 'Q', 'R']
                    long_width_cols = ['E', 'J', 'M',
                                       'O', 'P', 'S', 'T', 'W', 'X', 'Y']
                for idx, row in enumerate(worksheets[i].rows):
                    for cell in row:
                        if get_column_letter(cell.col_idx) in short_width_cols:
                            if idx == 0:
                                cell.alignment = vertical_alignment
                            else:
                                cell.alignment = center_alignment

                for dec in worksheets[i].columns:
                    col_letter = get_column_letter(index)
                    if col_letter in short_width_cols:
                        worksheets[i].column_dimensions[col_letter].width = short_width
                    elif col_letter in long_width_cols:
                        worksheets[i].column_dimensions[col_letter].width = long_width
                    else:
                        worksheets[i].column_dimensions[col_letter].width = normal_width
                    index = index + 1

    # set size field to Number format in "PACKAGES" and "RESOURCES"
    # size field location:
    # PACKAGES - AU
    # RESOURCES - AA
    # size field location for "reorder" flagged
    # PACKAGES - L
    # RESOURCES - I

    package_ws = wb['PACKAGES']
    package_col_letter = 'L'
    if not reorder:
        package_col_letter = 'AU'

    for index, cell in enumerate(package_ws[package_col_letter]):
        # Skip the header row
        if not index == 0 and cell.value:
            try:
                cell.value = int(cell.value)
            except:
                continue

    resource_ws = wb['RESOURCES']
    resource_col_letter = 'I'
    if not reorder:
        resource_col_letter = 'AA'
    for index, cell in enumerate(resource_ws[resource_col_letter]):
        # Skip the header row
        if not index == 0 and cell.value:
            try:
                cell.value = int(cell.value)
            except:
                continue

    # Reorder the worksheet in the following order:
    if not reorder:
        sheet_order = ['CHANGELOG', 'PACKAGES', 'RESOURCES',
                       'DEPENDENCIES', 'RELATIONS', 'LAYERS', 'ERRORS']
    else:
        wb.remove(wb.active)
        sheet_order = ['PACKAGES', 'RESOURCES',
                       'DEPENDENCIES', 'RELATIONS', 'LAYERS', 'ERRORS']

    # Get all the current sheetnames
    # sheetsname = wb.sheetnames

    # Create a new list of sheets in the desired order
    sorted_sheets = []
    for sheet_name in sheet_order:
        try:
            sorted_sheets.append(wb[sheet_name])
        except:
            # Pass if sheet name is not found
            pass

    # Reorder the sheets in the workbook
    wb._sheets = sorted_sheets

    return wb


def create_scio_report_bom(data, d2d_packages_data):
    '''
    Given a destination path and data from the input csv, write out the
    report XLSX file that contains the analysis data formatted to nexb
    report BOM standards.
    '''
    side = openpyxl.styles.borders.Side(border_style='thin', color='000000')
    border = openpyxl.styles.borders.Border(
        top=side, bottom=side, left=side, right=side)

    header_font = openpyxl.styles.fonts.Font(name='Calibri', size=9, b=True)
    bom_font = openpyxl.styles.fonts.Font(name='Calibri', size=9)

    about_file_color = openpyxl.styles.fills.PatternFill(
        'solid', fgColor='FFFFCC')
    report_color = openpyxl.styles.fills.PatternFill('solid', fgColor='D0F6FE')
    license_action_color = openpyxl.styles.fills.PatternFill(
        'solid', fgColor='FDE9D9')

    content_alignment = openpyxl.styles.alignment.Alignment(
        vertical='center', wrapText=True)
    vertical_alignment = Alignment(textRotation=90, horizontal="center")
    center_alignment = Alignment(horizontal="center", vertical='center')

    # Header formatting
    header_height = 82.00
    short_width = 4.00
    normal_width = 15.00
    long_width = 50.00
    super_long_width = 120.00
    readme_columnA_width = 25.30
    readme_columnB_width = 99.20

    wb = openpyxl.Workbook()
    readme_ws = wb.active

    readme_ws.title = "README"

    ws = wb.create_sheet('BOM')
    # The d2d_package_data is a list of list, so to check if the list contains
    # an empty list, we can check the first element of the list to see if it's
    # empty or not
    if d2d_packages_data[0]:
        d2d_ws = wb.create_sheet('D2D-packages')

    # Create README worksheet
    # opening the source excel file
    readme_file_location = os.path.normpath(os.path.join(
        os.path.dirname(__file__), '../reporting/files/cs_report_readme.xlsx'))

    readme_wb = load_workbook(readme_file_location)
    original_readme = readme_wb.worksheets[0]

    for row in original_readme.rows:
        for cell in row:
            new_cell = readme_ws.cell(
                row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    readme_index = 1
    for dec in readme_ws.columns:
        col_letter = get_column_letter(readme_index)
        if col_letter == 'A':
            readme_ws.column_dimensions[col_letter].width = readme_columnA_width
        elif col_letter == 'B':
            readme_ws.column_dimensions[col_letter].width = readme_columnB_width
        readme_index = readme_index + 1

    # Merge the Column A and B at row 44 and row 52
    readme_ws.merge_cells('A44:B44')
    readme_ws.merge_cells('A52:B52')

    # Write the main BOM worksheet
    for row in data:
        xlsx_row = []
        for idx, cell in enumerate(row):
            # The index is used to determine the color theme for the column
            try:
                xlsx_cell = openpyxl.cell.cell.Cell(worksheet=ws, value=cell)
            except:
                # ToDo: temp fix
                # Leave blank if the cell has encoding issue or otherwise
                # openpyxml will throw AttributionError
                xlsx_cell = openpyxl.cell.cell.Cell(worksheet=ws, value='')
            xlsx_cell.border = border
            if row == data[0]:
                xlsx_cell.font = header_font
                if idx <= 6 or (idx >= 8 and idx <= 20) or (idx > 22 and idx <= 40):
                    xlsx_cell.fill = report_color
                elif idx == 7 or idx == 21 or idx == 22 or idx == 41:
                    xlsx_cell.fill = license_action_color
                elif idx >= 42 and idx <= 48:
                    xlsx_cell.fill = about_file_color
            else:
                xlsx_cell.font = bom_font
                xlsx_cell.alignment = content_alignment
            xlsx_row.append(xlsx_cell)
        ws.append(xlsx_row)

    if d2d_packages_data[0]:
        # Write the D2D-packages worksheet
        for row in d2d_packages_data:
            xlsx_row = []
            for idx, cell in enumerate(row):
                # The index is used to determine the color theme for the column
                try:
                    xlsx_cell = openpyxl.cell.cell.Cell(
                        worksheet=d2d_ws, value=cell)
                except:
                    # ToDo: temp fix
                    # Leave blank if the cell has encoding issue or otherwise
                    # openpyxml will throw AttributionError
                    xlsx_cell = openpyxl.cell.cell.Cell(
                        worksheet=d2d_ws, value='')
                xlsx_cell.border = border
                if row == d2d_packages_data[0]:
                    xlsx_cell.font = header_font
                    if idx <= 5:
                        xlsx_cell.fill = report_color
                else:
                    xlsx_cell.font = bom_font
                    xlsx_cell.alignment = content_alignment
                xlsx_row.append(xlsx_cell)
            d2d_ws.append(xlsx_row)

        # This is the fomatting for the BOM worksheet
        # Set the height of the first row
        ws.row_dimensions[1].height = header_height
        index = 1
        # Columns letter for the BOM sheet
        short_width_cols = ['F', 'I', 'J', 'K',
                            'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'AI']
        long_width_cols = ['C', 'T', 'Y']
        for idx, row in enumerate(ws.rows):
            for cell in row:
                if get_column_letter(cell.col_idx) in short_width_cols:
                    if idx == 0:
                        cell.alignment = vertical_alignment
                    else:
                        cell.alignment = center_alignment

        for dec in ws.columns:
            col_letter = get_column_letter(index)
            if col_letter in short_width_cols:
                ws.column_dimensions[col_letter].width = short_width
            elif col_letter in long_width_cols:
                ws.column_dimensions[col_letter].width = long_width
            else:
                ws.column_dimensions[col_letter].width = normal_width
            index = index + 1

        # This is the formatting for the D2D-packages worksheet
        d2d_index = 1
        for dec in d2d_ws.columns:
            col_letter = get_column_letter(d2d_index)
            if col_letter == 'F':
                d2d_ws.column_dimensions[col_letter].width = super_long_width
            else:
                d2d_ws.column_dimensions[col_letter].width = normal_width
            d2d_index = d2d_index + 1

    return wb


def create_cs_report_bom(data, d2d_packages_data):
    '''
    Return a workbook with a 'BOM' and 'D2D-packages' worksheets
    '''
    side = openpyxl.styles.borders.Side(border_style='thin', color='000000')
    border = openpyxl.styles.borders.Border(
        top=side, bottom=side, left=side, right=side)

    header_font = openpyxl.styles.fonts.Font(name='Calibri', size=9, b=True)
    bom_font = openpyxl.styles.fonts.Font(name='Calibri', size=9)

    lcat_color = openpyxl.styles.fills.PatternFill('solid', fgColor='FFE599')
    report_color = openpyxl.styles.fills.PatternFill('solid', fgColor='D0F6FE')
    license_action_color = openpyxl.styles.fills.PatternFill(
        'solid', fgColor='F4CCCC')
    new_fields_color = openpyxl.styles.fills.PatternFill(
        'solid', fgColor='E2EFDA')
    about_fields_color = openpyxl.styles.fills.PatternFill(
        'solid', fgColor='FFFFCC')

    content_alignment = openpyxl.styles.alignment.Alignment(
        vertical='center', wrapText=True)
    vertical_alignment = Alignment(textRotation=90, horizontal="center")
    center_alignment = Alignment(horizontal="center", vertical='center')

    # Header formatting
    header_height = 82.00
    short_width = 4.00
    normal_width = 15.00
    long_width = 50.00
    super_long_width = 120.00

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"
    d2d_ws = wb.create_sheet("D2D-packages")

    # Write the main BOM worksheet
    for row in data:
        xlsx_row = []
        for idx, cell in enumerate(row):
            # The index is used to determine the color theme for the column
            try:
                xlsx_cell = openpyxl.cell.cell.Cell(worksheet=ws, value=cell)
            except:
                # ToDo: temp fix
                # Leave blank if the cell has encoding issue or otherwise
                # openpyxml will throw AttributionError
                xlsx_cell = openpyxl.cell.cell.Cell(worksheet=ws, value='')
            xlsx_cell.border = border
            if row == data[0]:
                xlsx_cell.font = header_font
                if idx == 6 or idx == 7 or idx == 12 or idx == 13:
                    xlsx_cell.fill = lcat_color
                elif idx == 8 or idx == 22 or idx == 23:
                    xlsx_cell.fill = license_action_color
                elif idx >= 42 and idx <= 44:
                    xlsx_cell.fill = new_fields_color
                elif idx >= 46:
                    xlsx_cell.fill = about_fields_color
                else:
                    xlsx_cell.fill = report_color
            else:
                xlsx_cell.font = bom_font
                xlsx_cell.alignment = content_alignment
            xlsx_row.append(xlsx_cell)
        ws.append(xlsx_row)

    # Write the D2D-packages worksheet
    for row in d2d_packages_data:
        xlsx_row = []
        for idx, cell in enumerate(row):
            # The index is used to determine the color theme for the column
            try:
                xlsx_cell = openpyxl.cell.cell.Cell(
                    worksheet=d2d_ws, value=cell)
            except:
                # ToDo: temp fix
                # Leave blank if the cell has encoding issue or otherwise
                # openpyxml will throw AttributionError
                xlsx_cell = openpyxl.cell.cell.Cell(worksheet=d2d_ws, value='')
            xlsx_cell.border = border
            if row == d2d_packages_data[0]:
                xlsx_cell.font = header_font
                if idx <= 5:
                    xlsx_cell.fill = report_color
            else:
                xlsx_cell.font = bom_font
                xlsx_cell.alignment = content_alignment
            xlsx_row.append(xlsx_cell)
        d2d_ws.append(xlsx_row)

    # This is the fomatting for the BOM worksheet
    # Set the height of the first row
    ws.row_dimensions[1].height = header_height
    index = 1
    # Columns letter for the BOM sheet
    short_width_cols = ['G', 'J', 'K', 'L',
                        'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'AJ']
    long_width_cols = ['D', 'U', 'Z', 'AA', 'AH', 'AU']
    for idx, row in enumerate(ws.rows):
        for cell in row:
            if get_column_letter(cell.col_idx) in short_width_cols:
                if idx == 0:
                    cell.alignment = vertical_alignment
                else:
                    cell.alignment = center_alignment

    for dec in ws.columns:
        col_letter = get_column_letter(index)
        if col_letter in short_width_cols:
            ws.column_dimensions[col_letter].width = short_width
        elif col_letter in long_width_cols:
            ws.column_dimensions[col_letter].width = long_width
        else:
            ws.column_dimensions[col_letter].width = normal_width
        index = index + 1

    # This is the formatting for the D2D-packages worksheet
    d2d_index = 1
    for dec in d2d_ws.columns:
        col_letter = get_column_letter(d2d_index)
        if col_letter == 'F':
            d2d_ws.column_dimensions[col_letter].width = super_long_width
        else:
            d2d_ws.column_dimensions[col_letter].width = normal_width
        d2d_index = d2d_index + 1

    # Freeze the top row
    wb["BOM"].freeze_panes = "A2"
    wb["D2D-packages"].freeze_panes = "A2"

    return wb


def create_scan2inv_format(data, wb, worksheet_name=None):
    '''
    Create the formatting for a scan2inv output .xlsx file.
    '''
    side = openpyxl.styles.borders.Side(border_style='thin', color='000000')
    border = openpyxl.styles.borders.Border(
        top=side, bottom=side, left=side, right=side)

    header_font = openpyxl.styles.fonts.Font(name='Calibri', size=10, b=True)
    header_color = openpyxl.styles.fills.PatternFill('solid', fgColor='c0fefd')
    header_alignment = openpyxl.styles.alignment.Alignment(vertical='bottom')

    content_font = openpyxl.styles.fonts.Font(name='Calibri', size=10)
    content_alignment = openpyxl.styles.alignment.Alignment(
        vertical='center', wrapText=False)

    if not worksheet_name:
        worksheet_name = 'BOM'
    # Put the "RESOURCES" sheet after the "PACKAGES"
    if "PACKAGES" in wb.sheetnames:
        packages_index = wb.index(wb['PACKAGES'])
        ws = wb.create_sheet(worksheet_name, packages_index + 1)
    else:
        ws = wb.create_sheet(worksheet_name)

    from openpyxl.styles import Alignment

    force_int_index_list = []
    hide_fields_list = ['md5', 'sha1', 'sha256', 'sha512', 'is_binary',
                        'is_text', 'is_archive', 'is_media', 'is_legal',
                        'is_manifest', 'is_readme', 'is_top_level',
                        'is_key_file']
    hide_fields_col_letter = []
    path_field_idx = ''
    path_field_col_letter = ''
    max_len_path = 0

    for row in data:
        xlsx_row = []
        for idx, cell in enumerate(row):
            idx += 1

            if idx in force_int_index_list and cell:
                # Force the cell value to be an integer
                xlsx_cell = openpyxl.cell.cell.Cell(
                    worksheet=ws, value=int(cell))
            else:
                xlsx_cell = openpyxl.cell.cell.Cell(worksheet=ws, value=cell)
            xlsx_cell.border = border
            if row == data[0]:
                if cell in hide_fields_list:
                    hide_fields_col_letter.append(get_column_letter(idx))
                if cell == 'analysis_priority':
                    force_int_index_list.append(idx)
                    xlsx_cell.font = openpyxl.styles.fonts.Font(
                        name='Calibri', size=10, b=True, italic=True)
                    xlsx_cell.fill = openpyxl.styles.fills.PatternFill(
                        'solid', fgColor='fff2cc')
                    xlsx_cell.alignment = Alignment(textRotation=90)
                    xlsx_cell.alignment = openpyxl.styles.alignment.Alignment(
                        horizontal='center', wrapText=False, textRotation=90)
                elif cell == 'file_category' or cell == 'file_subcategory':
                    xlsx_cell.font = openpyxl.styles.fonts.Font(
                        name='Calibri', size=10, b=True, italic=True)
                    xlsx_cell.fill = openpyxl.styles.fills.PatternFill(
                        'solid', fgColor='fff2cc')
                    xlsx_cell.alignment = openpyxl.styles.alignment.Alignment(
                        horizontal='left', wrapText=True)
                elif cell == 'Party':
                    xlsx_cell.font = openpyxl.styles.fonts.Font(
                        name='Calibri', size=10, b=True, italic=True)
                    xlsx_cell.fill = openpyxl.styles.fills.PatternFill(
                        'solid', fgColor='f2f2f2')
                    xlsx_cell.alignment = Alignment(textRotation=90)
                    xlsx_cell.alignment = openpyxl.styles.alignment.Alignment(
                        horizontal='center', wrapText=False, textRotation=90)
                elif cell == 'Status' or cell == 'Notes' or cell == 'ToDo':
                    xlsx_cell.font = openpyxl.styles.fonts.Font(
                        name='Calibri', size=10, b=True, italic=True)
                    xlsx_cell.fill = openpyxl.styles.fills.PatternFill(
                        'solid', fgColor='f2f2f2')
                    xlsx_cell.alignment = openpyxl.styles.alignment.Alignment(
                        horizontal='left', wrapText=True)
                else:
                    if cell == 'size':
                        force_int_index_list.append(idx)
                    elif cell == 'path':
                        path_field_idx = idx
                        path_field_col_letter = get_column_letter(idx)
                    xlsx_cell.font = header_font
                    xlsx_cell.fill = header_color
            else:
                # Calculate the length of the path field
                if idx == path_field_idx:
                    if len(str(cell)) > max_len_path:
                        max_len_path = len(str(cell))

                xlsx_cell.font = content_font
                xlsx_cell.alignment = content_alignment
            xlsx_row.append(xlsx_cell)
        ws.append(xlsx_row)

    # Hide the columns
    for col_letter in hide_fields_col_letter:
        ws.column_dimensions[col_letter].hidden = True

    # Expand the path field
    ws.column_dimensions[path_field_col_letter].width = max_len_path

    # Freeze the header row for all worksheets
    for worksheet in wb.worksheets:
        # Freeze the first row
        worksheet.freeze_panes = worksheet["A2"]
        # Update font and alignment for all worksheets
        # No need to work on 'RESOURCES' as it's been hendled
        if not worksheet.title == 'RESOURCES':
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    if cell.row == 1:
                        cell.font = header_font
                        cell.alignment = header_alignment
                    else:
                        cell.font = content_font
                        cell.alignment = content_alignment

    return wb


def remove_version_char(value):
    """
    Return a version value without a pre-prended 'v' character, as is the
    typical formatting convention at nexB.
    """
    if not value:
        return
    elif not value.lower().startswith('v'):
        return value
    else:
        value = value.strip('v')
        value = value.strip()
        return value


def get_column(worksheet, column_character):
    """
    Return a Tuple containg all the cells of a particular column
    in an excel Worksheet, chosen by the column character (A, B, AA, etc)
    """
    for column in worksheet.iter_cols():
        for cell in column:
            if cell.column_letter == column_character:
                return column


def count_headers(workbooks):
    """
    Return a sorted dictionary where {key: value} is {header_name: occurance #}
    in a given list of openpyxl workbooks.
    """
    header_names = []
    for workbook in workbooks:
        for sheet in workbook:
            headers = get_headers(sheet)
            if headers:
                header_names.extend(headers)

    header_totals = dict(Counter(header_names))
    return sorted(header_totals.items(), key=operator.itemgetter(1), reverse=True)


def count_sheets(workbooks):
    """
    Return a sorted dictionary where {key: value} is {worksheet_name: occurance #}
    in a given list of openpyxl workbooks.
    """
    sheet_names = []
    for workbook in workbooks:
        sheet_names.extend(workbook.sheetnames)

    sheet_totals = dict(Counter(sheet_names))
    return sorted(sheet_totals.items(), key=operator.itemgetter(1), reverse=True)


def get_headers(worksheet):
    """
    Returns a list of column headers that are in the given worksheet.
    """
    # by convention, column headers are on the first row of a worksheet.
    for row in worksheet.rows:
        headers = [curate_value(cell.value) for cell in row]
        return headers


def get_header_fields_and_index_dict(worksheet):
    """
    Return a dictionary list with header field as the key and the column index as the value
    """
    header_dict = {}
    for cell in worksheet[1]:
        header_dict[cell.value] = cell.column_letter
    return header_dict


def curate_value(value):
    ""'''""
    Returns a normalized value in utf-8 format. Handles datetime objects, numbers,
    strings, etc.
    '''
    if not value:
        return u''
    value = str(value).strip()
    return value  # .encode('utf-8')


def has_macros(workbook):
    """
    Returns a list of cell locations if the OpenPYXL workbook contains excel
    macros or functions.
    """
    cell_locations = []
    for worksheet in workbook:
        for row in worksheet.iter_rows():
            cell_locations.extend(
                [cell.coordinate for cell in row if cell.data_type == 'f'])

    return cell_locations


# FIXME: handle files in a better way; too much resource overhead required
#        using this method.
def load_xlsx_files(files):
    """
    Open multiple .xlsx files and return a dict in the form
    {'filename', workbook_object}
    """
    workbooks = {}
    for f in files:
        name = os.path.basename(f)
        workbook = (name, openpyxl.load_workbook(f))
        workbooks.update([workbook])
    return workbooks


def get_xlsx_paths(location):
    """
    Return a list of valid Excel file paths from a given 'location' input.
    """
    files = []
    if os.path.isfile(location) and location.endswith('.xlsx'):
        files.append(location)
    else:
        for top, _, filenames in os.walk(location):
            for filename in filenames:
                if filename.endswith('.xlsx'):
                    xlsx_file = os.path.join(top, filename)
                    files.append(xlsx_file)
    return files


def get_lic_key_and_coord(expression_info):
    # Get the license_key and coordinate
    lic_key_coord_dict = {}
    for le in expression_info:
        if le.key in lic_key_coord_dict:
            lic_key_coord_dict[le.key].append(le.coord)
        else:
            lic_key_coord_dict[le.key] = list([le.coord])
    return lic_key_coord_dict


def get_value_and_coord(bom, col_letter):
    """
    Given a BOM and the column letter, return a dictionary list
    with key as the value of the particular cell and the value
    as the coordinate index
    """
    ws = bom.active
    dict = {}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter != col_letter:
                continue
            if cell.value:
                if cell.value not in dict:
                    dict[cell.value] = [cell.coordinate]
                else:
                    dict[cell.value].append(cell.coordinate)
    return dict


def format_dict_data_for_xlsx_output(data_list):
    """
    This function is the convert the list of dictionary
    [{'Resource': 'xxx', 'package__type': 'npm', 'package__namespace': ''}]
    to [['Resource', 'package__type', 'package__namespace'], ['xxx', 'npm', '']]
    as this is the format needed for create_xlsx_output.
    """
    formatted_list = []
    key_list = []
    # Get all the dictionary keys from the list
    for data_dict in data_list:
        for key in data_dict:
            if key not in key_list:
                key_list.append(key)

    headers = key_list
    formatted_list.append(headers)
    for entry in data_list:
        entry_list = []
        have_content = False
        for header in headers:
            if header not in entry:
                entry[header] = ""
            # Special treatment to prevent excel to treat this cell as formula
            if entry[header] and str(entry[header]).startswith('='):
                special_treatment = "'" + entry[header]
                entry_list.append(special_treatment)
            else:
                entry_list.append(entry[header])
        # Prevent empty entry
        for entry in entry_list:
            if entry:
                have_content = True
                break
        if have_content:
            formatted_list.append(entry_list)
    return formatted_list


def create_xlsx_output(destination, data, input_bom=None, input_ws=None):
    """
    data is a list of list which the first list element is a list of header row follow
    by other list of content
    For instance,
    data = [['Resource', 'name', license], ['/tmp/', 'tmp', 'gpl-2.0']]
    """
    header_font = openpyxl.styles.fonts.Font(
        name='Calibri', size=10, bold=True)
    bom_font = openpyxl.styles.fonts.Font(name='Calibri', size=10)
    content_alignment = openpyxl.styles.alignment.Alignment(
        vertical='center', wrapText=True)
    report_color = openpyxl.styles.fills.PatternFill('solid', fgColor='C0FEFD')

    output_workbook = openpyxl.Workbook()
    if input_bom:
        # Remove the default sheet created in the new workbook
        output_workbook.remove(output_workbook.active)

        # Iterate through all sheets in the input workbook
        for sheet_name in input_bom.sheetnames:
            input_sheet = input_bom[sheet_name]

            # Copy the sheet to the output workbook
            output_sheet = output_workbook.create_sheet(title=sheet_name)
            # Freeze the top row
            output_sheet.freeze_panes = "A2"

            # Copy the column widths
            for col in input_sheet.columns:
                col_letter = col[0].column_letter
                output_sheet.column_dimensions[col_letter].width = input_sheet.column_dimensions[col_letter].width

            for row in input_sheet.iter_rows():
                for cell in row:
                    new_cell = output_sheet[cell.coordinate]
                    new_cell.value = cell.value
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = copy(cell.number_format)
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)
            if sheet_name == input_ws.title:
                last_cell_col = None
                # Get the last cell column
                for cell in input_ws[1]:
                    if cell.value:
                        last_cell_col = cell.column
                for row_idx, data_row in enumerate(data):
                    r_idx = row_idx + 1
                    for col_idx, lcat_data in enumerate(data_row):
                        next_column = openpyxl.utils.get_column_letter(
                            last_cell_col + col_idx + 1)
                        next_coordinate = f"{next_column}{r_idx}"
                        output_sheet[next_coordinate].value = lcat_data
                        if r_idx == 1:
                            output_sheet[next_coordinate].fill = report_color
                            output_sheet[next_coordinate].font = header_font
                        else:
                            output_sheet[next_coordinate].font = bom_font
                            output_sheet[next_coordinate].alignment = content_alignment
    else:
        output_sheet = output_workbook.active

        # Freeze the top row
        output_sheet.freeze_panes = "A2"

        for row in data:
            output_sheet.append(row)

    output_workbook.save(destination)


def create_multiple_worksheets_xlsx_output(destination, ws_name_list, data_list):
    """
    data_list is a list that contains multiple data that is a list of list which
    the first list element is a list of header row follow by other list of
    content
    For instance,
    ws_name_list = ['Worksheet1', 'Worksheet2']
    data_list = [[['Resource', 'name'], ['/tmp/ws1', 'ws1']], [['Resource', 'name'], ['/tmp/ws2', 'ws2']]]
    """
    assert len(ws_name_list) == len(data_list)
    new_bom = openpyxl.Workbook()
    # Remove the default sheet created by openpyxl
    new_bom.remove(new_bom.active)
    for index, ws_name in enumerate(ws_name_list):

        # write new bom
        ws_name = str(index)
        ws_name = new_bom.create_sheet(title=ws_name_list[index])

        # Write content to the worksheet
        data = data_list[index]
        for row in data:
            ws_name.append(row)
        # Freeze the top row
        ws_name.freeze_panes = 'A2'

    new_bom.save(destination)


def get_sheetname_from_xlsx(input):
    """
    Get all the sheetnames from the input
    """
    input_bom = load_workbook(input)
    return input_bom.sheetnames


def get_data_from_xlsx(input, sheet_name=None):
    """
    This function read the input xlsx file and return a
    list of dictionary of its content and the headers list
    """
    input_bom = load_workbook(input)
    if sheet_name:
        ws = input_bom[sheet_name]
    else:
        ws = input_bom.active

    rows = []
    results = []
    headers = []

    # Create tuple of column names
    column_names = []
    for cell in ws[1]:
        column_names.append(cell.value)
    header = tuple(column_names)

    # Use rows, created above, to create a list of dictionaries and set the min_row to 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(header, row)))

    # Handle situation where only header row exist and no value
    if not rows:
        header_row = {}
        for h in header:
            header_row[h] = ''
            headers.append(h)
        results.append(header_row)
    else:
        # Replace None values with empty string to avoid openpyxl errors -- requires a new dict
        for row in rows:
            # Append "if key" to avoid creating empty column
            row = {key: '' if value is None else value for (
                key, value) in row.items() if key}
            results.append(row)

        headers = [cell.value for cell in ws[1] if cell.value]

    return results, headers


def write_to_json(data, output):
    """
    Given a list of dictioanry data.
    Write JSON format output
    """
    with open(output, 'wb') as outfile:
        json.dump(list(data), outfile)


def write_to_csv(data_list, output):
    """
    Given a list of dictioanry data.
    Write CSV format output
    """
    key_list = []
    # Get all the dictionary keys from the list
    for data_dict in data_list:
        for key in data_dict:
            if key not in key_list:
                key_list.append(key)

    with open(output, 'w', encoding='utf-8-sig', newline='') as csvfile:
        fieldnames = key_list
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for item in data_list:
            # Prevent empty row
            have_content = False
            for key in item:
                if item[key]:
                    have_content = True
                    break
            if have_content:
                writer.writerow(item)


def write_to_xlsx(data_list, output):
    """
    Given a list of dictioanry data.
    Write XLSX format output
    """
    formatted_data = format_dict_data_for_xlsx_output(data_list)
    create_xlsx_output(output, formatted_data)
